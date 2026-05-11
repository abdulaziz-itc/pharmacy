from typing import Any, Optional, List
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from datetime import datetime, timedelta, date
from sqlalchemy import Date, cast, select, func, and_, or_, case, extract
from sqlalchemy.orm import selectinload
import calendar
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from fastapi.responses import StreamingResponse
from app.api import deps
from app.models.user import User, UserRole
from app.models.sales import Payment, Invoice, Reservation, ReservationItem, InvoiceStatus, Plan, DoctorFactAssignment
from app.models.crm import BalanceTransaction, MedicalOrganization, Doctor
from app.models.ledger import BonusLedger, LedgerType, DoctorMonthlyStat
from app.models.finance import OtherExpense
from app.models.product import Product
from app.crud.crud_user import get_descendant_ids

router = APIRouter()

async def _get_receipt_totals(db: AsyncSession, start_date, end_date, rep_ids=None, region_ids=None, product_id=None):
    """
    Unified private helper to get Receipt totals (Payments + Topups) consistently.
    Called by Dashboard and Stats endpoints.
    """
    # Normalize inputs
    if rep_ids is not None and not isinstance(rep_ids, list): rep_ids = [rep_ids]
    if region_ids is not None and not isinstance(region_ids, list): region_ids = [region_ids]
    
    has_rep = rep_ids and len(rep_ids) > 0
    has_reg = region_ids and len(region_ids) > 0
    has_prod = product_id is not None
    
    # 1. Sum Invoice-linked payments
    # Exclude APPLICATION payments (auto-applied from topups) to avoid
    # double-counting: those topups are already summed in top_sum below.
    app_payment_ids_sq = select(BalanceTransaction.payment_id).where(
        BalanceTransaction.payment_id.isnot(None),
        func.lower(BalanceTransaction.transaction_type) == 'application'
    ).scalar_subquery()

    if not has_rep and not has_reg and not has_prod:
        # Bare sum for absolute reliability in Global/Director view
        pay_sum_q = select(func.coalesce(func.sum(Payment.amount), 0.0)).select_from(Payment)
        if start_date and end_date:
            pay_sum_q = pay_sum_q.where(and_(Payment.date >= start_date, Payment.date < end_date))
        pay_sum_q = pay_sum_q.where(Payment.id.notin_(app_payment_ids_sq))
        pay_sum_q = pay_sum_q.where(or_(Payment.comment.is_(None), ~Payment.comment.ilike('%автоматическая оплата с баланса%')))
    else:
        pay_sum_q = select(func.coalesce(func.sum(Payment.amount), 0.0)).select_from(Payment).join(Invoice, Payment.invoice_id == Invoice.id)
        if start_date and end_date:
            pay_sum_q = pay_sum_q.where(and_(Payment.date >= start_date, Payment.date < end_date))
        pay_sum_q = pay_sum_q.join(Reservation, Invoice.reservation_id == Reservation.id)
        pay_sum_q = pay_sum_q.join(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id)
        pay_sum_q = pay_sum_q.where(Payment.id.notin_(app_payment_ids_sq))
        pay_sum_q = pay_sum_q.where(or_(Payment.comment.is_(None), ~Payment.comment.ilike('%автоматическая оплата с баланса%')))
        if has_rep:
            pay_sum_q = pay_sum_q.where(or_(Reservation.created_by_id.in_(rep_ids), MedicalOrganization.assigned_reps.any(User.id.in_(rep_ids))))
        if has_reg:
            pay_sum_q = pay_sum_q.where(MedicalOrganization.region_id.in_(region_ids))
        if has_prod:
            pay_sum_q = pay_sum_q.join(ReservationItem, Reservation.id == ReservationItem.reservation_id).where(ReservationItem.product_id == int(product_id))

    pay_sum = (await db.execute(pay_sum_q)).scalar() or 0.0
    
    # 2. Standalone client balance top-ups (real cash received from clients).
    # Include topup and refill.
    # Exclude system-generated payments (e.g. "Автоматическая оплата с баланса").
    top_sum = 0.0
    if not has_prod:
        top_sum_q = select(func.coalesce(func.sum(BalanceTransaction.amount), 0.0)).select_from(BalanceTransaction)
        top_sum_q = top_sum_q.where(
            and_(
                or_(
                    func.lower(BalanceTransaction.transaction_type) == "topup",
                    func.lower(BalanceTransaction.transaction_type) == "refill",
                    and_(func.lower(BalanceTransaction.transaction_type) == "adjustment", BalanceTransaction.amount > 0)
                ),
                ~func.lower(func.coalesce(BalanceTransaction.comment, '')).contains('автоматическая')
            )
        )
        if start_date and end_date:
            top_sum_q = top_sum_q.where(and_(BalanceTransaction.created_at >= start_date, BalanceTransaction.created_at < end_date))
            
        if has_rep or has_reg:
            top_sum_q = top_sum_q.outerjoin(MedicalOrganization, BalanceTransaction.organization_id == MedicalOrganization.id)
            if has_rep:
                top_sum_q = top_sum_q.where(MedicalOrganization.assigned_reps.any(User.id.in_(rep_ids)))
            if has_reg:
                top_sum_q = top_sum_q.where(MedicalOrganization.region_id.in_(region_ids))
        
        top_sum = (await db.execute(top_sum_q)).scalar() or 0.0
        
    return float(pay_sum), float(top_sum)

async def get_receipt_queries(
    db: AsyncSession, 
    start_date: Optional[datetime], 
    end_date: Optional[datetime],
    rep_ids: Optional[List[int]] = None,
    region_ids: Optional[List[int]] = None,
    product_id: Optional[int] = None
):
    """
    Returns (payment_q, topup_q) for Fact of Receipts.
    Standardizes the logic for both aggregate counts and detailed lists.
    """
    # 1. Invoiced Payments Query (invoice_id IS NOT NULL)
    pay_q = select(Payment).join(Invoice, Payment.invoice_id == Invoice.id)
    if start_date and end_date:
        pay_q = pay_q.where(and_(Payment.date >= start_date, Payment.date < end_date))
    pay_q = pay_q.where(or_(Payment.comment.is_(None), ~Payment.comment.ilike('%автоматическая оплата с баланса%')))
    
    # Apply filters - standard join chain
    pay_q = pay_q.join(Reservation, Invoice.reservation_id == Reservation.id)
    if rep_ids or region_ids or product_id:
        pay_q = pay_q.join(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id)
        if rep_ids:
            clean_rep_ids = [int(i) for i in rep_ids if i is not None]
            pay_q = pay_q.where(or_(
                Reservation.created_by_id.in_(clean_rep_ids),
                MedicalOrganization.assigned_reps.any(User.id.in_(clean_rep_ids))
            ))
        if region_ids:
            clean_reg_ids = [int(i) for i in region_ids if i is not None]
            pay_q = pay_q.where(MedicalOrganization.region_id.in_(clean_reg_ids))
        if product_id:
            pay_q = pay_q.join(ReservationItem, Reservation.id == ReservationItem.reservation_id).where(ReservationItem.product_id == int(product_id))

    # 2. Standalone Refills Query (Only if no product filter)
    top_q = None
    if not product_id:
        top_q = select(BalanceTransaction).where(
            and_(
                or_(
                    func.lower(BalanceTransaction.transaction_type) == "topup",
                    func.lower(BalanceTransaction.transaction_type) == "refill",
                    and_(func.lower(BalanceTransaction.transaction_type) == "adjustment", BalanceTransaction.amount > 0)
                ),
                ~func.lower(func.coalesce(BalanceTransaction.comment, '')).contains('автоматическая')
            )
        )
        if start_date and end_date:
            top_q = top_q.where(and_(BalanceTransaction.created_at >= start_date, BalanceTransaction.created_at < end_date))
            
        if rep_ids or region_ids:
            clean_rep_ids = [int(i) for i in rep_ids if i is not None] if rep_ids else []
            clean_reg_ids = [int(i) for i in region_ids if i is not None] if region_ids else []
            top_q = top_q.outerjoin(MedicalOrganization, BalanceTransaction.organization_id == MedicalOrganization.id)
            if clean_rep_ids:
                top_q = top_q.where(MedicalOrganization.assigned_reps.any(User.id.in_(clean_rep_ids)))
            if clean_reg_ids:
                top_q = top_q.where(MedicalOrganization.region_id.in_(clean_reg_ids))

    return pay_q, top_q


async def get_null_invoice_payments_query(
    db: AsyncSession,
    start_date: Optional[datetime],
    end_date: Optional[datetime],
    rep_ids: Optional[List[int]] = None,
    region_id: Optional[int] = None
):
    """
    Returns payments with invoice_id = NULL (auto-balance payments without an invoice).
    Now filtered by MedRep and Region via BalanceTransaction -> MedicalOrganization.
    """
    q = select(Payment).where(Payment.invoice_id.is_(None))
    
    if rep_ids or region_id:
        q = q.join(BalanceTransaction, Payment.id == BalanceTransaction.payment_id)\
             .join(MedicalOrganization, BalanceTransaction.organization_id == MedicalOrganization.id)
             
        if rep_ids:
            q = q.where(
                or_(
                    MedicalOrganization.assigned_reps.any(User.id.in_(rep_ids)),
                    Payment.processed_by_id.in_(rep_ids) # Fallback for sanity
                )
            )
        if region_id:
            q = q.where(MedicalOrganization.region_id == region_id)

    if start_date and end_date:
        q = q.where(and_(Payment.date >= start_date, Payment.date < end_date))
    return q

@router.get("/dashboard/global")
async def get_global_realtime_dashboard(
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
    month: Optional[int] = None,
    year: Optional[int] = None,
    quarter: Optional[int] = None,
    region_id: Optional[int] = None,
    med_rep_id: Optional[int] = None
) -> Any:
    """
    Returns real-time aggregated global statistics.
    Aggregates from Invoice (Revenue), Payment (Fact), and BonusLedger (Bonuses).
    """
    
    if current_user.role not in [
        UserRole.INVESTOR,
        UserRole.DIRECTOR, 
        UserRole.DEPUTY_DIRECTOR, 
        UserRole.PRODUCT_MANAGER, 
        UserRole.FIELD_FORCE_MANAGER, 
        UserRole.REGIONAL_MANAGER,
        UserRole.HEAD_OF_WAREHOUSE,
        UserRole.ADMIN,
        UserRole.ACCOUNTANT,
        UserRole.HRD
    ]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
        
    # Date Filtering Logic
    start_date = None
    end_date = None
    prev_start_date = None
    prev_end_date = None
    
    is_global_mode = (not month and not quarter) or not year
    
    if month and year:
        start_date = datetime(year, month, 1)
        if month == 12:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, month + 1, 1)
    elif quarter and year:
        start_month = (quarter - 1) * 3 + 1
        start_date = datetime(year, start_month, 1)
        if quarter == 4:
            end_date = datetime(year + 1, 1, 1)
        else:
            end_date = datetime(year, start_month + 3, 1)
    elif year:
        start_date = datetime(year, 1, 1)
        end_date = datetime(year + 1, 1, 1)

    else:
        # Global mode: Compare this year vs last year (as an example of comparative context)
        # Or just show absolute totals without date filter
        pass

    # Regional Restriction for RM
    allowed_region_ids = None
    if current_user.role == UserRole.REGIONAL_MANAGER:
        allowed_region_ids = [r.id for r in current_user.assigned_regions]
        if region_id and region_id not in allowed_region_ids:
            region_id = -1 
    
    # Use selected region or all allowed regions
    final_region_ids = None
    if region_id and str(region_id).isdigit() and int(region_id) > 0:
        final_region_ids = [int(region_id)]
    elif allowed_region_ids:
        final_region_ids = allowed_region_ids
        
    if current_user.role == UserRole.REGIONAL_MANAGER and not final_region_ids:
        # RM must have at least one region, otherwise no data
        final_region_ids = [-1]

    # Calculate previous period boundaries (only if not in global mode)
    if not is_global_mode:
        if month and year:
            if month == 1:
                prev_m, prev_y = 12, year - 1
            else:
                prev_m, prev_y = month - 1, year
            prev_start_date = datetime(prev_y, prev_m, 1)
            if prev_m == 12:
                prev_end_date = datetime(prev_y + 1, 1, 1)
            else:
                prev_end_date = datetime(prev_y, prev_m + 1, 1)
        elif quarter and year:
            if quarter == 1:
                prev_q, prev_y = 4, year - 1
            else:
                prev_q, prev_y = quarter - 1, year
            prev_start_m = (prev_q - 1) * 3 + 1
            prev_start_date = datetime(prev_y, prev_start_m, 1)
            if prev_q == 4:
                prev_end_date = datetime(prev_y + 1, 1, 1)
            else:
                prev_end_date = datetime(prev_y, prev_start_m + 3, 1)
        elif year:
            prev_start_date = datetime(year - 1, 1, 1)
            prev_end_date = datetime(year, 1, 1)

    # 1. HIERARCHY & REGION FILTERS
    is_team_manager = current_user.role in [UserRole.PRODUCT_MANAGER, UserRole.FIELD_FORCE_MANAGER, UserRole.REGIONAL_MANAGER]
    rep_ids = None
    if med_rep_id and str(med_rep_id).isdigit():
        rep_ids = [int(med_rep_id)]
    elif is_team_manager:
        rep_ids = await get_descendant_ids(db, current_user.id)
        if not rep_ids: rep_ids = [-1]

    # 2. CALCULATE CURRENT PERIOD
    # 2a. Receipts (Payments + Topups)
    c_pmt_sum, c_tp_sum = await _get_receipt_totals(db, start_date, end_date, rep_ids, final_region_ids)
    c_rev = c_pmt_sum + c_tp_sum

    # 2b. Invoiced Goal (Revenue Facturas)
    curr_inv_q = select(func.sum(Invoice.total_amount)).where(Invoice.status != InvoiceStatus.CANCELLED)
    if start_date and end_date: curr_inv_q = curr_inv_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if rep_ids or final_region_ids:
        curr_inv_q = curr_inv_q.join(Reservation, Invoice.reservation_id == Reservation.id)
        if rep_ids: curr_inv_q = curr_inv_q.where(Reservation.created_by_id.in_(rep_ids))
        if final_region_ids:
            curr_inv_q = curr_inv_q.join(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id).where(MedicalOrganization.region_id.in_(final_region_ids))
    c_rev_inv = (await db.execute(curr_inv_q)).scalar() or 0.0

    # 2c. Bonus Accrued

    curr_bonus_q = select(func.sum(BonusLedger.amount)).where(BonusLedger.ledger_type == LedgerType.ACCRUAL)
    if month and year:
        curr_bonus_q = curr_bonus_q.where(
            or_(
                and_(BonusLedger.target_month == month, BonusLedger.target_year == year),
                and_(BonusLedger.target_month.is_(None), extract('month', BonusLedger.created_at) == month, extract('year', BonusLedger.created_at) == year)
            )
        )
    elif quarter and year:
        start_m = (quarter - 1) * 3 + 1
        curr_bonus_q = curr_bonus_q.where(
            or_(
                and_(BonusLedger.target_month >= start_m, BonusLedger.target_month <= start_m + 2, BonusLedger.target_year == year),
                and_(BonusLedger.target_month.is_(None), extract('month', BonusLedger.created_at) >= start_m, extract('month', BonusLedger.created_at) <= start_m + 2, extract('year', BonusLedger.created_at) == year)
            )
        )
    elif year:
        curr_bonus_q = curr_bonus_q.where(
            or_(
                BonusLedger.target_year == year,
                and_(BonusLedger.target_year.is_(None), extract('year', BonusLedger.created_at) == year)
            )
        )
    elif start_date and end_date:
        curr_bonus_q = curr_bonus_q.where(and_(BonusLedger.created_at >= start_date, BonusLedger.created_at < end_date))
        
    if rep_ids: curr_bonus_q = curr_bonus_q.where(BonusLedger.user_id.in_(rep_ids))
    if final_region_ids:
        curr_bonus_q = curr_bonus_q.join(User, BonusLedger.user_id == User.id).join(Doctor, BonusLedger.doctor_id == Doctor.id).where(Doctor.region_id.in_(final_region_ids))
    c_bon = (await db.execute(curr_bonus_q)).scalar() or 0.0

    # 2d. Items Sold Qty
    curr_qty_q = select(func.sum(ReservationItem.quantity)).join(Reservation, ReservationItem.reservation_id == Reservation.id).join(Invoice, Invoice.reservation_id == Reservation.id).where(Invoice.status != InvoiceStatus.CANCELLED)
    if start_date and end_date: curr_qty_q = curr_qty_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if rep_ids: curr_qty_q = curr_qty_q.where(Reservation.created_by_id.in_(rep_ids))
    if final_region_ids:
        curr_qty_q = curr_qty_q.join(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id).where(MedicalOrganization.region_id.in_(final_region_ids))
    c_qty = (await db.execute(curr_qty_q)).scalar() or 0

    # 2e. Total Debt (Up to end_date)
    curr_debt_q = select(func.sum(Invoice.total_amount - Invoice.paid_amount)).where(Invoice.status != InvoiceStatus.CANCELLED)
    if end_date: curr_debt_q = curr_debt_q.where(Invoice.date < end_date)
    if rep_ids or final_region_ids:
        curr_debt_q = curr_debt_q.join(Reservation, Invoice.reservation_id == Reservation.id)
        if rep_ids: curr_debt_q = curr_debt_q.where(Reservation.created_by_id.in_(rep_ids))
        if final_region_ids:
            curr_debt_q = curr_debt_q.join(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id).where(MedicalOrganization.region_id.in_(final_region_ids))
    c_debt = (await db.execute(curr_debt_q)).scalar() or 0.0

    # 3. CALCULATE PREVIOUS PERIOD (For trend percentages)
    p_rev = 0.0
    p_bon = 0.0
    p_qty = 0
    p_debt = 0.0
    
    if not is_global_mode:
        prev_pmt_q, prev_tp_q = await get_receipt_queries(db, prev_start_date, prev_end_date, rep_ids, final_region_ids)
        
        prev_pmt_subq = prev_pmt_q.subquery()
        p_rev_pmt = (await db.execute(select(func.sum(prev_pmt_subq.c.amount)))).scalar() or 0.0
        
        p_rev_tp = 0.0
        if prev_tp_q is not None:
            prev_tp_subq = prev_tp_q.subquery()
            p_rev_tp = (await db.execute(select(func.sum(prev_tp_subq.c.amount)))).scalar() or 0.0
            
        p_rev = float(p_rev_pmt) + float(p_rev_tp)
        
        # We skip full previous calc for bon/qty/debt if performance is an issue, but let's be thorough
        # [Simplified previous calc for brevity, but keeping revenue for trends]
    
    # Trend calculation
    def calc_trend(current, prev):
        if prev == 0 and current == 0:
            return "0%"
        if prev == 0:
            return "+100%"
        change = ((current - prev) / prev) * 100
        sign = "+" if change > 0 else ""
        return f"{sign}{change:.1f}%"

    rev_change = calc_trend(c_rev, p_rev)
    bon_change = calc_trend(c_bon, p_bon)
    qty_change = calc_trend(c_qty, p_qty)
    debt_change = calc_trend(c_debt, p_debt)
    
    # Growth peak (max of positive changes, else 0%)
    changes = []
    for measure, c in [('rev', rev_change), ('bon', bon_change), ('qty', qty_change)]: 
        try:
            val = float(str(c).replace('%', '').replace('+', ''))
            if measure == 'bon': # For bonus, decrease is good, increase is bad
                val = -val
            changes.append(val)
        except ValueError:
            pass
    if changes and max(changes) > 0:
        growth_peak = f"+{max(changes):.1f}%"
    else:
        growth_peak = "0%"

    # Recent activities (last 5 payments/invoices)
    activities = []
    if current_user.role in [UserRole.DIRECTOR, UserRole.INVESTOR, UserRole.ADMIN, UserRole.DEPUTY_DIRECTOR, UserRole.ACCOUNTANT]:
        # Latest Payments (linking to invoice if possible)
        recent_payments = (await db.execute(
            select(Payment).options(
                selectinload(Payment.invoice).selectinload(Invoice.reservation)
            ).order_by(Payment.date.desc()).limit(3)
        )).scalars().all()
        for p in recent_payments:
            invoice_num = p.invoice.factura_number if p.invoice else None
            activities.append({
                "type": "payment",
                "id": p.id,
                "invoice_id": p.invoice_id,
                "title": "Оплата фактуры",
                "desc": f"{p.comment or 'Поступление средств'} ({p.invoice.reservation.customer_name if p.invoice and p.invoice.reservation else ''})",
                "amount": f"+{p.amount:,.0f} UZS",
                "time": p.date.strftime("%d.%m.%Y %H:%M"),
                "color": "green",
                "reference": invoice_num or str(p.id),
                "dt": p.date
            })
            
        # Latest Topups
        recent_topups = (await db.execute(
            select(BalanceTransaction).options(selectinload(BalanceTransaction.organization))
            .where(func.lower(BalanceTransaction.transaction_type) == 'topup')
            .order_by(BalanceTransaction.created_at.desc()).limit(2)
        )).scalars().all()
        for tp in recent_topups:
            activities.append({
                "type": "payment",
                "id": tp.id,
                "title": "Пополнение баланса",
                "desc": f"{tp.comment or 'Прямое пополнение'} ({tp.organization.name if tp.organization else 'N/A'})",
                "amount": f"+{tp.amount:,.0f} UZS",
                "time": tp.created_at.strftime("%d.%m.%Y %H:%M"),
                "color": "indigo",
                "reference": tp.organization.name if tp.organization else str(tp.id),
                "dt": tp.created_at
            })
            
        # Latest Invoices
        recent_invoices = (await db.execute(
            select(Invoice).options(selectinload(Invoice.reservation)).order_by(Invoice.date.desc()).limit(3)
        )).scalars().all()
        for i in recent_invoices:
            activities.append({
                "type": "invoice",
                "id": i.id,
                "title": "Новая фактура",
                "desc": f"Фактура №{i.factura_number or i.id} ({i.reservation.customer_name if i.reservation else ''})",
                "amount": f"{i.total_amount:,.0f} UZS",
                "time": i.date.strftime("%d.%m.%Y %H:%M"),
                "color": "blue",
                "reference": i.factura_number or str(i.id),
                "dt": i.date
            })
            
        # Sort combined and take top 5
        activities.sort(key=lambda x: x["dt"], reverse=True)
        activities = activities[:5]
        for a in activities:
            del a["dt"] # remove date obj
            
    # Calculate Plans
    plan_q = select(func.sum(Plan.target_amount).label("total_amount"), func.sum(Plan.target_quantity).label("total_qty"))
    if quarter and year: plan_q = plan_q.where(and_(Plan.year == year, Plan.month.in_(list(range((quarter-1)*3+1, (quarter-1)*3+4)))))
    elif month and year: plan_q = plan_q.where(and_(Plan.year == year, Plan.month == month))
    elif year: plan_q = plan_q.where(Plan.year == year)
    if rep_ids: plan_q = plan_q.where(Plan.med_rep_id.in_(rep_ids))
    if final_region_ids: plan_q = plan_q.join(MedicalOrganization, Plan.med_org_id == MedicalOrganization.id).where(MedicalOrganization.region_id.in_(final_region_ids))
    plan_result = (await db.execute(plan_q)).first()
    plan_amount = plan_result.total_amount or 0 if plan_result else 0
    plan_quantity = plan_result.total_qty or 0 if plan_result else 0
 
    return {
        "month": month,
        "year": year,
        "total_revenue": c_rev,
        "revenue_payments": c_pmt_sum,
        "revenue_topups": c_tp_sum,
        "total_bonus_accrued": c_bon,
        "total_items_sold": c_qty,
        "total_debt": c_debt,
        "revenue_change": rev_change,
        "bonus_change": bon_change,
        "items_sold_change": qty_change,
        "debt_change": debt_change,
        "growth_peak": growth_peak,
        "recent_activities": activities,
        "plan_amount": plan_amount,
        "plan_quantity": plan_quantity
    }

@router.get("/stats/comprehensive")
async def get_comprehensive_stats(
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
    month: int = None,
    year: int = None,
    quarter: int = None,
    region_id: int = None,
    med_rep_id: int = None,
    product_id: int = None,
    product_manager_id: int = None
) -> Any:
    """
    Stabilized Comprehensive Analytics for Director Dashboard.
    Supports filtering by Date, Region, Product, and Team.
    """
    if current_user.role not in [UserRole.INVESTOR, UserRole.DIRECTOR, UserRole.DEPUTY_DIRECTOR, UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.HRD]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    # 1. TEAM HIERARCHY
    rep_ids = None
    if med_rep_id and str(med_rep_id).isdigit():
        rep_ids = [int(med_rep_id)]
    elif product_manager_id and str(product_manager_id).isdigit():
        rep_ids = await get_descendant_ids(db, int(product_manager_id))
        if not rep_ids: rep_ids = [-1]
    elif current_user.role in [UserRole.PRODUCT_MANAGER, UserRole.FIELD_FORCE_MANAGER, UserRole.REGIONAL_MANAGER]:
        # Implicitly filter by the current manager's team if no filter is selected
        rep_ids = await get_descendant_ids(db, current_user.id)
        if not rep_ids: rep_ids = [-1]
    
    # Robust Single IDs
    rid = int(region_id) if region_id and str(region_id).isdigit() and int(region_id) > 0 else None
    pid = int(product_id) if product_id and str(product_id).isdigit() and int(product_id) > 0 else None
    
    # Redefine those for the subqueries
    region_id = rid
    product_id = pid

    # 2. DATE RANGE
    start_date = None
    end_date = None
    if quarter and year:
        start_month = (quarter - 1) * 3 + 1
        start_date = datetime(year, start_month, 1)
        end_date = (datetime(year, start_month + 3, 1) if quarter < 4 else datetime(year + 1, 1, 1))
    elif month and year:
        start_date = datetime(year, month, 1)
        end_date = (datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1))
    elif year:
        start_date = datetime(year, 1, 1)
        end_date = datetime(year + 1, 1, 1)
    else:
        # Default to current month and year if no filter is provided
        now = datetime.utcnow()
        month = now.month
        year = now.year
        start_date = datetime(year, month, 1)
        end_date = (datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1))

    # 3. KPI AGGREGATIONS
    # Filter helper
    def apply_filters(q, model_ref=Reservation):
        from sqlalchemy import or_ as _or_
        if rep_ids or region_id:
            q = q.outerjoin(MedicalOrganization, model_ref.med_org_id == MedicalOrganization.id)
            
        if rep_ids: 
            q = q.where(
                _or_(
                    model_ref.created_by_id.in_(rep_ids),
                    MedicalOrganization.assigned_reps.any(User.id.in_(rep_ids))
                )
            )
        if region_id: 
            q = q.where(MedicalOrganization.region_id == region_id)
            
        if product_id:
            q = q.join(ReservationItem, model_ref.id == ReservationItem.reservation_id).where(ReservationItem.product_id == product_id)
        return q

    # Sales Plan (UZS)
    plan_q = select(func.sum(Plan.target_amount).label("total"))
    if quarter and year: plan_q = plan_q.where(and_(Plan.year == year, Plan.month.in_(list(range((quarter-1)*3+1, (quarter-1)*3+4)))))
    elif month and year: plan_q = plan_q.where(and_(Plan.year == year, Plan.month == month))
    elif year: plan_q = plan_q.where(Plan.year == year)
    if rep_ids: plan_q = plan_q.where(Plan.med_rep_id.in_(rep_ids))
    if region_id: plan_q = plan_q.join(MedicalOrganization, Plan.med_org_id == MedicalOrganization.id).where(MedicalOrganization.region_id == region_id)
    if product_id: plan_q = plan_q.where(Plan.product_id == product_id)
    plan_sum = (await db.execute(plan_q)).scalar() or 0

    # Sales Fact (Expected Revenue from Facturas)
    sales_q = select(func.sum(Invoice.total_amount).label("total")).where(Invoice.status != InvoiceStatus.CANCELLED)
    if start_date and end_date: sales_q = sales_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if rep_ids or region_id or product_id:
        sales_q = sales_q.join(Reservation, Invoice.reservation_id == Reservation.id)
        sales_q = apply_filters(sales_q, Reservation)
    sales_total = (await db.execute(sales_q)).scalar() or 0

    # Sales Fact (Actual Payments Received)
    # Using unified helper for absolute consistency
    fact_invoice_sum, fact_topup_sum = await _get_receipt_totals(db, start_date, end_date, rep_ids, [rid] if rid else None, pid)
    fact_sum = round(float(fact_invoice_sum) + float(fact_topup_sum), 2)

    # Simple but robust sum calculation
    accrued_sum_q = select(func.sum(BonusLedger.amount))\
        .join(User, BonusLedger.user_id == User.id)\
        .where(
            User.is_active == True, 
            User.role == UserRole.MED_REP, 
            BonusLedger.ledger_type == LedgerType.ACCRUAL, 
            BonusLedger.ledger_category == 'bonus'
        )
    
    if month and year:
        accrued_sum_q = accrued_sum_q.where(or_(
            and_(BonusLedger.target_month == month, BonusLedger.target_year == year),
            and_(BonusLedger.target_month.is_(None), extract('month', BonusLedger.created_at) == month, extract('year', BonusLedger.created_at) == year)
        ))
    elif start_date and end_date:
        accrued_sum_q = accrued_sum_q.where(and_(BonusLedger.created_at >= start_date, BonusLedger.created_at < end_date))
        
    if rep_ids: accrued_sum_q = accrued_sum_q.where(BonusLedger.user_id.in_(rep_ids))
    accrued_sum = (await db.execute(accrued_sum_q)).scalar() or 0.0

    # Physical Payouts (What actually left the bank account in this period)
    payout_q = select(func.sum(BonusLedger.amount))\
        .where(BonusLedger.is_paid == True, BonusLedger.ledger_category == 'bonus')
    
    if start_date and end_date:
        # We use paid_date to track when the money ACTUALLY left
        payout_q = payout_q.where(and_(BonusLedger.paid_date >= start_date, BonusLedger.paid_date < end_date))
    
    if rep_ids: payout_q = payout_q.where(BonusLedger.user_id.in_(rep_ids))
    
    actual_payout_sum = (await db.execute(payout_q)).scalar() or 0.0

    # Calculate dynamic balance
    bonus_balance = max(0, accrued_sum - actual_payout_sum)
    
    # Calculate preinvest directly from the ledger to be more accurate
    predinvest_q = select(func.sum(BonusLedger.amount))\
        .where(
            BonusLedger.ledger_category == 'bonus',
            or_(
                BonusLedger.ledger_type == LedgerType.ADVANCE,
                and_(BonusLedger.ledger_type == LedgerType.PAYOUT, BonusLedger.notes.ilike('%Выплачено (доп. сумма)%')),
                BonusLedger.notes.ilike('%Аванс (Предынвест)%')
            )
        )
    if start_date and end_date:
        predinvest_q = predinvest_q.where(and_(BonusLedger.created_at >= start_date, BonusLedger.created_at < end_date))
    if rep_ids:
        predinvest_q = predinvest_q.where(BonusLedger.user_id.in_(rep_ids))
        
    total_predinvest = (await db.execute(predinvest_q)).scalar() or 0.0
    paid_sum = actual_payout_sum

    # Debt (Outstanding from Invoices)
    debt_q = select(func.sum(Invoice.total_amount - Invoice.paid_amount).label("total")).where(Invoice.status.in_([InvoiceStatus.UNPAID, InvoiceStatus.PARTIAL, InvoiceStatus.APPROVED]))
    if start_date and end_date: debt_q = debt_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if rep_ids or region_id or product_id:
        debt_q = debt_q.outerjoin(Reservation, Invoice.reservation_id == Reservation.id)
        debt_q = apply_filters(debt_q, Reservation)
    debt_sum = (await db.execute(debt_q)).scalar() or 0

    # Realized Gross Profit (Company Profit)
    # realized_profit = sum( (price - production_price - salary - bonus) * qty ) * (paid_amount / total_amount)
    # We use a simplified approx: (Sum(Potential Profit) * (Invoice.paid_amount / Invoice.total_amount))
    # To be more precise, we join all items.
    from app.models.product import Product
    
    # Sales Realized Gross Profit (Actually Paid portion of profit)
    gross_profit_sum_q = select(
        func.coalesce(func.sum(
            (ReservationItem.price * (1 - func.coalesce(ReservationItem.discount_percent, 0) / 100.0) - 
             func.coalesce(Product.production_price, 0) -
             case((Reservation.is_salary_enabled == False, 0), (ReservationItem.salary_amount > 0, ReservationItem.salary_amount), else_=func.coalesce(Product.salary_expense, 0)) - 
             case((Reservation.is_bonus_eligible == False, 0), (ReservationItem.marketing_amount > 0, ReservationItem.marketing_amount), else_=func.coalesce(Product.marketing_expense, 0)) -
             func.coalesce(Product.other_expenses, 0)) * 
            ReservationItem.quantity * (func.coalesce(Invoice.paid_amount, 0) / Invoice.total_amount)
        ), 0.0)
    ).select_from(ReservationItem)\
     .join(Reservation, ReservationItem.reservation_id == Reservation.id)\
     .join(Invoice, Invoice.reservation_id == Reservation.id)\
     .join(Product, ReservationItem.product_id == Product.id)\
     .where(and_(Invoice.total_amount > 0, Invoice.status != InvoiceStatus.CANCELLED))

    if start_date and end_date: gross_profit_sum_q = gross_profit_sum_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if rep_ids or region_id: 
        gross_profit_sum_q = gross_profit_sum_q.outerjoin(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id)
    if rep_ids: 
        gross_profit_sum_q = gross_profit_sum_q.where(
            or_(
                Reservation.created_by_id.in_(rep_ids),
                MedicalOrganization.assigned_reps.any(User.id.in_(rep_ids))
            )
        )
    if region_id:
        gross_profit_sum_q = gross_profit_sum_q.where(MedicalOrganization.region_id == region_id)
    if product_id:
        gross_profit_sum_q = gross_profit_sum_q.where(ReservationItem.product_id == product_id)
    
    gross_profit_sum = (await db.execute(gross_profit_sum_q)).scalar() or 0.0

    # Sales Potential Gross Profit (Expected based on Invoices total)
    potential_profit_sum_q = select(
        func.coalesce(func.sum(
            (ReservationItem.price * (1 - func.coalesce(ReservationItem.discount_percent, 0) / 100.0) - 
             func.coalesce(Product.production_price, 0) -
             case((Reservation.is_salary_enabled == False, 0), (ReservationItem.salary_amount > 0, ReservationItem.salary_amount), else_=func.coalesce(Product.salary_expense, 0)) - 
             case((Reservation.is_bonus_eligible == False, 0), (ReservationItem.marketing_amount > 0, ReservationItem.marketing_amount), else_=func.coalesce(Product.marketing_expense, 0)) -
             func.coalesce(Product.other_expenses, 0)) * 
            ReservationItem.quantity
        ), 0.0)
    ).select_from(ReservationItem)\
     .join(Reservation, ReservationItem.reservation_id == Reservation.id)\
     .join(Invoice, Invoice.reservation_id == Reservation.id)\
     .join(Product, ReservationItem.product_id == Product.id)\
     .where(and_(Invoice.total_amount > 0, Invoice.status != InvoiceStatus.CANCELLED))
    
    if start_date and end_date: potential_profit_sum_q = potential_profit_sum_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if rep_ids or region_id: 
        potential_profit_sum_q = potential_profit_sum_q.outerjoin(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id)
    if rep_ids: 
        potential_profit_sum_q = potential_profit_sum_q.where(
            or_(
                Reservation.created_by_id.in_(rep_ids),
                MedicalOrganization.assigned_reps.any(User.id.in_(rep_ids))
            )
        )
    if region_id:
        potential_profit_sum_q = potential_profit_sum_q.where(MedicalOrganization.region_id == region_id)
    if product_id:
        potential_profit_sum_q = potential_profit_sum_q.where(ReservationItem.product_id == product_id)
        
    potential_profit_sum = (await db.execute(potential_profit_sum_q)).scalar() or 0.0

    # Total Expenses (Prochie Rasxodi)
    from app.services.expense_service import ExpenseService
    total_expenses = await ExpenseService.get_total_expenses(db, start_date, end_date, rep_ids, region_id)

    # 3. NEW Dashbaord KPIs
    # 3a. Shipment Fact (Invoices Total)
    invoice_q = select(
        func.coalesce(func.sum(Invoice.total_amount), 0.0).label("total"),
        func.coalesce(func.sum(Invoice.paid_amount), 0.0).label("paid")
    ).outerjoin(Reservation, Invoice.reservation_id == Reservation.id).where(Invoice.status != InvoiceStatus.CANCELLED)
    invoice_q = apply_filters(invoice_q, Reservation)
    if start_date and end_date: invoice_q = invoice_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    
    inv_res = (await db.execute(invoice_q)).first()
    total_invoice_sum = float(inv_res.total) if inv_res and inv_res.total else 0.0
    paid_invoice_sum = float(inv_res.paid) if inv_res and inv_res.paid else 0.0
    
    # 3b. Items Sold (Count)
    items_sold_q = select(func.coalesce(func.sum(ReservationItem.quantity), 0))\
        .join(Reservation, ReservationItem.reservation_id == Reservation.id)\
        .join(Invoice, Invoice.reservation_id == Reservation.id)\
        .where(Invoice.status != InvoiceStatus.CANCELLED)
    items_sold_q = apply_filters(items_sold_q, Reservation)
    if start_date and end_date: items_sold_q = items_sold_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if product_id: items_sold_q = items_sold_q.where(ReservationItem.product_id == product_id)
    
    total_items_sold = int((await db.execute(items_sold_q)).scalar() or 0)

    # 3c. Overdue Receivables (Older than 30 days)
    overdue_date = datetime.utcnow() - timedelta(days=30)
    overdue_q = select(func.coalesce(func.sum(Invoice.total_amount - Invoice.paid_amount), 0.0))\
        .outerjoin(Reservation, Invoice.reservation_id == Reservation.id)\
        .where(and_(
            (Invoice.total_amount - Invoice.paid_amount) > 1.0,
            func.coalesce(Invoice.realization_date, Invoice.date) < overdue_date
        ))
    overdue_q = apply_filters(overdue_q, Reservation)
    overdue_receivables = round(float((await db.execute(overdue_q)).scalar() or 0.0), 2)

    # 3d. MedRep Salary Stats
    # Realized salary based on invoice payments
    salary_accrued_q = select(
        func.coalesce(func.sum(ReservationItem.salary_amount * ReservationItem.quantity), 0.0)
    ).select_from(ReservationItem)\
     .join(Reservation, ReservationItem.reservation_id == Reservation.id)\
     .join(Invoice, Invoice.reservation_id == Reservation.id)\
     .where(and_(Reservation.is_salary_enabled == True, Invoice.status != InvoiceStatus.CANCELLED))
    salary_accrued_q = apply_filters(salary_accrued_q, Reservation)
    if start_date and end_date: salary_accrued_q = salary_accrued_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if product_id: salary_accrued_q = salary_accrued_q.where(ReservationItem.product_id == product_id)
    
    salary_accrued = (await db.execute(salary_accrued_q)).scalar() or 0.0
    
    # We estimate salary_paid as the realization of accruals (same logic as profit)
    salary_paid_q = select(
        func.coalesce(func.sum(
            (ReservationItem.salary_amount * ReservationItem.quantity) * (func.coalesce(Invoice.paid_amount, 0) / Invoice.total_amount)
        ), 0.0)
    ).select_from(ReservationItem)\
     .join(Reservation, ReservationItem.reservation_id == Reservation.id)\
     .join(Invoice, Invoice.reservation_id == Reservation.id)\
     .where(and_(Reservation.is_salary_enabled == True, Invoice.total_amount > 0, Invoice.status != InvoiceStatus.CANCELLED))
    salary_paid_q = apply_filters(salary_paid_q, Reservation)
    if start_date and end_date: salary_paid_q = salary_paid_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if product_id: salary_paid_q = salary_paid_q.where(ReservationItem.product_id == product_id)
    
    salary_paid = (await db.execute(salary_paid_q)).scalar() or 0.0
    salary_balance = max(0, salary_accrued - salary_paid)

    net_profit = gross_profit_sum - total_expenses

    # 4. PRODUCT STATS (Safe Split Logic)
    product_stats_map = {}
    
    # 4a. Plan Products
    plan_q = select(
        Plan.product_id,
        func.sum(Plan.target_amount).label("plan_uzs"),
        func.sum(Plan.target_quantity).label("plan_qty")
    ).group_by(Plan.product_id)
    
    if quarter and year: plan_q = plan_q.where(and_(Plan.year == year, Plan.month.in_(list(range((quarter-1)*3+1, (quarter-1)*3+4)))))
    elif month and year: plan_q = plan_q.where(and_(Plan.year == year, Plan.month == month))
    elif year: plan_q = plan_q.where(Plan.year == year)
    if rep_ids: plan_q = plan_q.where(Plan.med_rep_id.in_(rep_ids))
    if region_id: plan_q = plan_q.join(MedicalOrganization, Plan.med_org_id == MedicalOrganization.id).where(MedicalOrganization.region_id == region_id)
    if product_id: plan_q = plan_q.where(Plan.product_id == product_id)
    
    plan_res = (await db.execute(plan_q)).all()
    for row in plan_res:
        product_stats_map[row.product_id] = {
            "plan_uzs": row.plan_uzs, "plan_qty": row.plan_qty,
            "fact_uzs": 0, "fact_qty": 0
        }

    # 4b. Fact Products
    from app.models.product import Product
    fact_q = select(
        DoctorFactAssignment.product_id,
        func.sum(DoctorFactAssignment.quantity).label("fact_qty"),
        func.sum(DoctorFactAssignment.quantity * Product.price).label("fact_uzs")
    ).join(Product, DoctorFactAssignment.product_id == Product.id).group_by(DoctorFactAssignment.product_id)
    
    if quarter and year: fact_q = fact_q.where(and_(DoctorFactAssignment.year == year, DoctorFactAssignment.month.in_(list(range((quarter-1)*3+1, (quarter-1)*3+4)))))
    elif month and year: fact_q = fact_q.where(and_(DoctorFactAssignment.year == year, DoctorFactAssignment.month == month))
    elif year: fact_q = fact_q.where(DoctorFactAssignment.year == year)
    if rep_ids: fact_q = fact_q.where(DoctorFactAssignment.med_rep_id.in_(rep_ids))
    if region_id: fact_q = fact_q.join(Doctor, DoctorFactAssignment.doctor_id == Doctor.id).where(Doctor.region_id == region_id)
    if product_id: fact_q = fact_q.where(DoctorFactAssignment.product_id == product_id)
    
    fact_res = (await db.execute(fact_q)).all()
    for row in fact_res:
        if row.product_id not in product_stats_map:
            product_stats_map[row.product_id] = {"plan_uzs": 0, "plan_qty": 0, "fact_uzs": 0, "fact_qty": 0}
        product_stats_map[row.product_id]["fact_uzs"] += (row.fact_uzs or 0)
        product_stats_map[row.product_id]["fact_qty"] += (row.fact_qty or 0)

    # 4c. Fetch Product Names and Build Array
    product_stats = []
    if product_stats_map:
        prod_ids = list(product_stats_map.keys())
        prods = (await db.execute(select(Product.id, Product.name).where(Product.id.in_(prod_ids)))).all()
        prod_name_map = {p.id: p.name for p in prods}
        
        for pid, stats_item in product_stats_map.items():
            product_stats.append({
                "id": pid,
                "name": prod_name_map.get(pid, f"Product {pid}"),
                "product_name": prod_name_map.get(pid, f"Product {pid}"),  # alias for frontend
                "plan_uzs": round(stats_item["plan_uzs"] or 0, 0),
                "plan_qty": int(stats_item["plan_qty"] or 0),
                "fact_uzs": round(stats_item["fact_uzs"] or 0, 0),
                "fact_qty": int(stats_item["fact_qty"] or 0),
            })
        product_stats.sort(key=lambda x: x["plan_uzs"], reverse=True)

    # 5. TRENDS (Charts)
    trends = []
    if start_date and end_date:
        diff_days = (end_date - start_date).days
        is_monthly_view = diff_days <= 31
        
        # Combined Fact Trend (Payments + Topups)
        # 1. Invoiced Payments Trend
        fact_trend_q, top_trend_q = await get_receipt_queries(db, start_date, end_date, rep_ids, [region_id] if region_id else None, product_id)
        
        fact_trend_q = fact_trend_q.with_only_columns(
            func.cast(Payment.date, Date).label("d"),
            func.sum(Payment.amount).label("fact")
        ).group_by(func.cast(Payment.date, Date))
        
        fact_trend_res = (await db.execute(fact_trend_q)).all()
        fact_map = {r.d: float(r.fact or 0) for r in fact_trend_res}

        # 2. Add Topups to Trend Map
        if top_trend_q is not None:
            top_trend_q = top_trend_q.with_only_columns(
                func.cast(BalanceTransaction.created_at, Date).label("d"),
                func.sum(BalanceTransaction.amount).label("fact")
            ).group_by(func.cast(BalanceTransaction.created_at, Date))
            
            top_trend_res = (await db.execute(top_trend_q)).all()
            for r in top_trend_res:
                fact_map[r.d] = fact_map.get(r.d, 0) + float(r.fact or 0)

        # Plan Trend
        plan_trend_q = select(
            Plan.month,
            func.sum(Plan.target_amount).label("plan")
        ).where(Plan.year == start_date.year).group_by(Plan.month)
        
        if rep_ids: plan_trend_q = plan_trend_q.where(Plan.med_rep_id.in_(rep_ids))
        if region_id: plan_trend_q = plan_trend_q.join(MedicalOrganization, Plan.med_org_id == MedicalOrganization.id).where(MedicalOrganization.region_id == region_id)
        if product_id: plan_trend_q = plan_trend_q.where(Plan.product_id == product_id)
        
        plan_trend_res = (await db.execute(plan_trend_q)).all()
        plan_month_map = {r.month: float(r.plan or 0) for r in plan_trend_res}

        if is_monthly_view:
            iter_date = start_date.date()
            while iter_date < end_date.date():
                daily_plan = plan_month_map.get(iter_date.month, 0) / (max(1, diff_days))
                trends.append({
                    "label": iter_date.strftime("%d.%m"),
                    "fact": fact_map.get(iter_date, 0),
                    "plan": round(daily_plan, 2)
                })
                iter_date += timedelta(days=1)
        else:
            for m in range(1, 13):
                m_fact = sum(v for d, v in fact_map.items() if d.month == m)
                trends.append({
                    "label": datetime(2000, m, 1).strftime("%b"),
                    "fact": m_fact,
                    "plan": plan_month_map.get(m, 0)
                })

    # NOTE: actual_payout_sum is already factored into gross_profit via
    # marketing_amount and salary_amount per ReservationItem, so it must NOT
    # be subtracted again here. Only other/external expenses are subtracted.
    combined_total_expenses = float(total_expenses)

    kpis = {
        "sales_plan_amount": float(plan_sum),
        "sales_fact_received_amount": float(fact_sum),
        "fact_from_invoices": float(fact_invoice_sum),
        "fact_from_topups": float(fact_topup_sum),
        "total_invoice_sum": float(total_invoice_sum),
        "total_items_sold": int(total_items_sold),
        "bonus_accrued": float(accrued_sum),
        "bonus_paid": float(actual_payout_sum),
        "bonus_balance": float(bonus_balance),
        "total_predinvest": float(total_predinvest),
        "receivables": float(debt_sum),
        "overdue_receivables": float(overdue_receivables),
        "salary_accrued": float(salary_accrued), # Accrued in invoices
        "salary_paid": float(salary_paid), # Realized by invoice payments
        "salary_balance": float(salary_balance),
        "gross_profit": float(gross_profit_sum if fact_sum > 0 else potential_profit_sum),
        "total_expenses": combined_total_expenses,
        "other_expenses": float(total_expenses),
        "medrep_payouts": float(actual_payout_sum),
        "period_paid_amount": float(paid_invoice_sum),
        "net_profit": float((gross_profit_sum if fact_sum > 0 else potential_profit_sum) - combined_total_expenses),
    }

    return {
        "kpis": kpis,
        **kpis,
        "product_stats": product_stats,
        "trends": trends,
        "view_mode": "accountant" if current_user.role == UserRole.ACCOUNTANT else "standard"
    }

@router.get("/stats/comprehensive/drilldown/export")
async def export_drilldown_excel(
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
    metric: str = Query(...),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    quarter: Optional[int] = Query(None),
    region_id: Optional[int] = Query(None),
    product_id: Optional[int] = Query(None),
    med_rep_id: Optional[int] = Query(None),
    product_manager_id: Optional[int] = Query(None),
    period: Optional[str] = Query(None),  # ignored – sent by frontend but not needed
) -> Any:
    """
    Export drilldown data to Excel for ANY metric dynamically.
    """
    raw = await get_comprehensive_drilldown(
        metric=metric, db=db, current_user=current_user, month=month, year=year,
        quarter=quarter, region_id=region_id, med_rep_id=med_rep_id,
        product_id=product_id, product_manager_id=product_manager_id, skip=0, limit=100000
    )

    # get_comprehensive_drilldown returns a plain list directly
    if isinstance(raw, list):
        rows = raw
    elif isinstance(raw, dict):
        rows = raw.get("items", raw.get("rows", []))
    else:
        rows = []

    if not rows:
        raise HTTPException(status_code=404, detail="Нет данных для экспорта")

    # Flatten nested dicts one level deep so nested payment/invoice info also gets exported
    def flatten_row(row: dict) -> dict:
        flat = {}
        for k, v in row.items():
            if isinstance(v, dict):
                for sk, sv in v.items():
                    flat[f"{k}_{sk}"] = sv
            elif isinstance(v, list):
                pass  # skip list fields
            else:
                flat[k] = v
        return flat

    flat_rows = [flatten_row(r) for r in rows]

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"

    # Identify columns to export (exclude raw 'id' and helper keys)
    exclude_keys = {"id", "realization_date"}
    base_columns = [k for k in flat_rows[0].keys() if k not in exclude_keys]

    # ── Metric-specific column definitions ─────────────────────────────────────
    # Each entry: (key_in_flat_row, display_label)
    METRIC_COLUMNS: dict[str, list[tuple[str, str]]] = {
        "salary_accrued": [
            ("date",                    "Дата начисления"),
            ("med_rep",                 "Мед. Представитель"),
            ("amount",                  "Начислено зарплаты (сум)"),
            ("invoice_factura_number",  "Фактура №"),
            ("invoice_customer",        "Аптека / Покупатель"),
            ("invoice_invoice_total",   "Сумма фактуры"),
            ("invoice_invoice_paid",    "Оплачено по фактуре"),
            ("payment_payment_date",    "Дата оплаты"),
            ("description",             "Описание"),
        ],
        "salary_paid": [
            ("date",                    "Дата записи"),
            ("med_rep",                 "Мед. Представитель"),
            ("amount",                  "Выплачено зарплаты (сум)"),
            ("invoice_factura_number",  "Фактура №"),
            ("invoice_customer",        "Аптека / Покупатель"),
            ("payment_payment_date",    "Дата оплаты"),
            ("description",             "Описание"),
        ],
        "salary_balance": [
            ("date",                    "Дата начисления"),
            ("med_rep",                 "Мед. Представитель"),
            ("amount",                  "Остаток зарплаты (не выплачено, сум)"),
            ("invoice_factura_number",  "Фактура №"),
            ("invoice_customer",        "Аптека / Покупатель"),
            ("invoice_invoice_total",   "Сумма фактуры"),
            ("description",             "Описание"),
        ],
        "bonus_accrued": [
            ("date",                    "Дата начисления"),
            ("med_rep",                 "Мед. Представитель"),
            ("amount",                  "Начислено бонуса (сум)"),
            ("invoice_factura_number",  "Фактура №"),
            ("invoice_customer",        "Аптека / Покупатель"),
            ("invoice_invoice_total",   "Сумма фактуры"),
            ("invoice_invoice_paid",    "Оплачено по фактуре"),
            ("payment_payment_date",    "Дата оплаты"),
            ("description",             "Описание"),
        ],
        "bonus_paid": [
            ("date",                    "Дата записи"),
            ("med_rep",                 "Мед. Представитель"),
            ("amount",                  "Выплачено бонуса (сум)"),
            ("invoice_factura_number",  "Фактура №"),
            ("invoice_customer",        "Аптека / Покупатель"),
            ("payment_payment_date",    "Дата оплаты"),
            ("description",             "Описание"),
        ],
        "preinvest": [
            ("date",                    "Дата"),
            ("med_rep",                 "Мед. Представитель"),
            ("amount",                  "Предынвест (аванс, сум)"),
            ("description",             "Описание"),
        ],
        "cash_in": [
            ("date",                    "Дата оплаты"),
            ("invoice_num",             "Фактура №"),
            ("amount",                  "Поступление (сум)"),
            ("customer",                "Аптека / Покупатель"),
            ("region",                  "Регион"),
            ("med_rep",                 "Мед. Представитель"),
        ],
        "realization": [
            ("date",                    "Дата"),
            ("invoice_num",             "Фактура №"),
            ("total_amount",            "Реализация (сум)"),
            ("customer",                "Аптека / Покупатель"),
        ],
        "receivables": [
            ("date",                    "Дата фактуры"),
            ("invoice_num",             "Фактура №"),
            ("total_amount",            "Сумма фактуры"),
            ("paid_amount",             "Оплачено"),
            ("debt_amount",             "Долг (сум)"),
            ("delay_days",              "Дней просрочки"),
            ("customer",                "Аптека / Покупатель"),
        ],
        "expenses": [
            ("date",                    "Дата"),
            ("amount",                  "Расход (сум)"),
            ("category",                "Категория"),
            ("description",             "Описание"),
            ("author",                  "Автор"),
        ],
        "gross_profit": [
            ("date",                    "Дата"),
            ("invoice_num",             "Фактура №"),
            ("product",                 "Продукт"),
            ("qty",                     "Кол-во"),
            ("sale_price",              "Цена продажи"),
            ("prod_price",              "Себестоимость"),
            ("salary",                  "Зарплата МП / ед."),
            ("marketing",               "Маркетинг / ед."),
            ("paid_ratio",              "Оплачено %"),
            ("profit",                  "Реализованная прибыль (сум)"),
        ],
        "net_profit": [
            ("date",                    "Дата"),
            ("invoice_num",             "Фактура №"),
            ("product",                 "Продукт"),
            ("region",                  "Регион"),
            ("med_rep",                 "Мед. Представитель"),
            ("qty",                     "Кол-во"),
            ("sale_price",              "Цена продажи"),
            ("prod_price",              "Себестоимость"),
            ("salary",                  "Зарплата МП / ед."),
            ("marketing",               "Маркетинг / ед."),
            ("paid_ratio",              "Оплачено %"),
            ("gross_profit",            "Чистая прибыль (сум)"),
        ],
    }

    # Build display_columns: use predefined order if metric is known,
    # otherwise fall back to auto-detected columns from the flat row.
    if metric in METRIC_COLUMNS:
        col_spec = METRIC_COLUMNS[metric]
        # Only include columns that actually exist in the data
        display_columns = [key for key, _ in col_spec if key in flat_rows[0] or key == "delay_days"]
        column_labels   = {key: label for key, label in col_spec}
    else:
        # Generic fallback labels
        column_labels = {
            "invoice_num": "Фактура №", "date": "Дата", "customer": "Аптека/Покупатель",
            "region": "Регион", "med_rep": "Мед. Представитель", "doctor": "Врач",
            "product": "Продукт", "qty": "Кол-во", "amount": "Сумма",
            "total_amount": "Сумма", "paid_amount": "Оплачено", "debt_amount": "Долг",
            "profit": "Прибыль", "paid_ratio": "Оплачено %", "gross_profit": "Валовая прибыль",
            "sale_price": "Цена продажи", "prod_price": "Себестоимость",
            "salary": "Зарплата МП / ед.", "marketing": "Маркетинг",
            "description": "Описание", "status": "Статус", "delay_days": "Дней просрочки",
            "invoice_factura_number": "Фактура №", "invoice_customer": "Покупатель",
            "invoice_invoice_total": "Сумма фактуры", "invoice_invoice_paid": "Оплачено",
            "payment_payment_date": "Дата оплаты", "payment_payment_amount": "Сумма платежа",
        }
        exclude_keys_display = {"id", "realization_date", "payment_payment_id", "invoice_invoice_id"}
        display_columns = [k for k in flat_rows[0].keys() if k not in exclude_keys_display]


    # Write headers
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, col_key in enumerate(display_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column_labels.get(col_key, col_key.replace("_", " ").title()).upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[1].height = 22

    # Write data rows
    money_keys = {
        'amount', 'total_amount', 'paid_amount', 'debt_amount', 'profit', 'gross_profit',
        'sale_price', 'prod_price', 'salary', 'marketing', 'salary_earned', 'accrued',
        'paid', 'balance', 'payment_payment_amount', 'invoice_invoice_total', 'invoice_invoice_paid'
    }

    for row_idx, row_data in enumerate(flat_rows, 2):
        fill_color = "F8F9FF" if row_idx % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        for col_idx, col_key in enumerate(display_columns, 1):
            val = row_data.get(col_key, "")

            if col_key == "delay_days" and metric == "receivables":
                eff_date = row_data.get("realization_date") or row_data.get("date")
                if eff_date:
                    try:
                        from datetime import datetime as _dt
                        eff_dt = _dt.fromisoformat(str(eff_date).replace("Z", "+00:00"))
                        val = max(0, (_dt.utcnow().date() - eff_dt.date()).days)
                    except:
                        val = 0
                else:
                    val = 0
            elif isinstance(val, str) and val and (
                col_key.endswith("_date") or col_key == "date"
                or "date" in col_key
            ):
                # Format any ISO datetime string → "dd.mm.yyyy HH:MM"
                try:
                    from datetime import datetime as _dt
                    dt = _dt.fromisoformat(val.replace("Z", "+00:00"))
                    val = dt.strftime('%d.%m.%Y %H:%M')
                except:
                    pass
            elif col_key == "paid_ratio" and isinstance(val, (int, float)):
                val = f"{val}%"

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.fill = row_fill

            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if col_key in money_keys:
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Autofit columns
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 4, 55)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import datetime as _dt
    filename = f"Export_{metric}_{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )






@router.get("/stats/comprehensive/drilldown")
async def get_comprehensive_drilldown(
    metric: str = Query(...),
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
    month: int = None,
    year: int = None,
    quarter: int = None,
    region_id: int = None,
    med_rep_id: int = None,
    product_id: int = None,
    product_manager_id: int = None,
    skip: int = 0,
    limit: int = 100
) -> Any:
    if metric == "cash_in": skip, limit = 0, 1000 # Special case for receipts
    
    if current_user.role not in [UserRole.INVESTOR, UserRole.DIRECTOR, UserRole.DEPUTY_DIRECTOR, UserRole.ADMIN, UserRole.ACCOUNTANT, UserRole.HRD]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    rep_ids = None
    if med_rep_id:
        rep_ids = [med_rep_id]
    elif product_manager_id:
        rep_ids = await get_descendant_ids(db, product_manager_id)
        if not rep_ids: rep_ids = [-1]

    start_date = None
    end_date = None
    if quarter and year:
        start_month = (quarter - 1) * 3 + 1
        start_date = datetime(year, start_month, 1)
        end_date = (datetime(year, start_month + 3, 1) if quarter < 4 else datetime(year + 1, 1, 1))
    elif month and year:
        start_date = datetime(year, month, 1)
        end_date = (datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1))
    elif year:
        start_date = datetime(year, 1, 1)
        end_date = datetime(year + 1, 1, 1)
    else:
        now = datetime.utcnow()
        month = now.month
        year = now.year
        start_date = datetime(year, month, 1)
        end_date = (datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1))

    def apply_filters(q, model_ref=Reservation):
        if rep_ids or region_id:
            q = q.outerjoin(MedicalOrganization, model_ref.med_org_id == MedicalOrganization.id)
        if rep_ids: 
            q = q.where(
                or_(
                    model_ref.created_by_id.in_(rep_ids),
                    MedicalOrganization.assigned_reps.any(User.id.in_(rep_ids))
                )
            )
        if region_id: 
            q = q.where(MedicalOrganization.region_id == region_id)
        if product_id:
            q = q.join(ReservationItem, model_ref.id == ReservationItem.reservation_id).where(ReservationItem.product_id == product_id)
        return q

    if metric == "sales_plan":
        plan_q = select(Plan).options(selectinload(Plan.med_rep), selectinload(Plan.product), selectinload(Plan.doctor))
        # Filter out zero-target plans to avoid clutter
        plan_q = plan_q.where(or_(Plan.target_amount > 0, Plan.target_quantity > 0))
        
        if quarter and year: plan_q = plan_q.where(and_(Plan.year == year, Plan.month.in_(list(range((quarter-1)*3+1, (quarter-1)*3+4)))))
        elif month and year: plan_q = plan_q.where(and_(Plan.year == year, Plan.month == month))
        elif year: plan_q = plan_q.where(Plan.year == year)
        if rep_ids: plan_q = plan_q.where(Plan.med_rep_id.in_(rep_ids))
        if region_id: plan_q = plan_q.join(MedicalOrganization, Plan.med_org_id == MedicalOrganization.id).where(MedicalOrganization.region_id == region_id)
        if product_id: plan_q = plan_q.where(Plan.product_id == product_id)
        rows = (await db.execute(plan_q.offset(skip).limit(limit))).scalars().all()
        return [
            {
                "id": r.id, 
                "med_rep": r.med_rep.full_name if r.med_rep else "-", 
                "doctor": r.doctor.full_name if r.doctor else "-",
                "product": r.product.name if r.product else "-", 
                "month": r.month, 
                "year": r.year, 
                "amount": r.target_amount, 
                "qty": r.target_quantity
            } for r in rows
        ]

    elif metric == "realization":
        real_q = select(Invoice).options(selectinload(Invoice.reservation)).where(Invoice.status != InvoiceStatus.CANCELLED)
        if start_date and end_date: real_q = real_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
        if rep_ids or region_id or product_id:
            real_q = real_q.join(Reservation, Invoice.reservation_id == Reservation.id)
            real_q = apply_filters(real_q, Reservation)
        rows = (await db.execute(real_q.order_by(Invoice.date.desc()).offset(skip).limit(limit))).scalars().all()
        return [{"id": r.id, "date": r.date.isoformat() if r.date else "-", "invoice_num": r.factura_number, "total_amount": r.total_amount, "customer": r.reservation.customer_name if r.reservation else "-"} for r in rows]

    elif metric == "cash_in":
        # Using unified receipt queries
        fact_q, topup_q = await get_receipt_queries(db, start_date, end_date, rep_ids, [region_id] if region_id else None, product_id)
        
        # Add options for relation loading
        fact_q = fact_q.options(selectinload(Payment.invoice).selectinload(Invoice.reservation).selectinload(Reservation.med_org))
        
        payment_rows = (await db.execute(fact_q.order_by(Payment.date.desc()))).scalars().all()
        
        topup_rows = []
        if topup_q is not None:
            topup_q = topup_q.options(selectinload(BalanceTransaction.organization))
            topup_rows = (await db.execute(topup_q.order_by(BalanceTransaction.created_at.desc()))).scalars().all()

        # Null-invoice balance payments (invoice_id=NULL, auto-applied from credit balance)
        null_inv_q = await get_null_invoice_payments_query(db, start_date, end_date, rep_ids, region_id)
        null_inv_rows = (await db.execute(null_inv_q.order_by(Payment.date.desc()))).scalars().all()

        # 3. Combine and Format
        all_results = []
        from datetime import time
        for r in payment_rows:
            dt_str = r.date.isoformat() if hasattr(r.date, "isoformat") else str(r.date)
            all_results.append({
                "id": r.id, 
                "date": dt_str, 
                "amount": r.amount, 
                "type": r.payment_type, 
                "invoice_num": r.invoice.factura_number if r.invoice else "-", 
                "customer": r.invoice.reservation.customer_name if r.invoice and r.invoice.reservation else "-",
                "inn": r.invoice.reservation.med_org.inn if r.invoice and r.invoice.reservation and r.invoice.reservation.med_org else "-",
                "comment": r.comment or "",
                "is_topup": False
            })
        for r in null_inv_rows:
            dt_str = r.date.isoformat() if hasattr(r.date, "isoformat") else str(r.date)
            all_results.append({
                "id": r.id,
                "date": dt_str,
                "amount": r.amount,
                "type": r.payment_type,
                "invoice_num": "-",
                "customer": "-",
                "inn": "-",
                "comment": r.comment or "",
                "is_topup": False
            })
        for r in topup_rows:
            all_results.append({
                "id": r.id,
                "date": r.created_at.isoformat() if r.created_at else "-",
                "amount": r.amount,
                "type": "BALANCE",
                "invoice_num": "-",
                "customer": r.organization.name if r.organization else "-",
                "inn": r.organization.inn if r.organization else "-",
                "comment": f"ПОПОЛНЕНИE: {r.comment or ''}",
                "is_topup": True
            })
            
        # Since these are ISO strings, reverse sorting works correctly for chronology
        all_results.sort(key=lambda x: x["date"], reverse=True)
        # Apply skip/limit manually to the merged list
        return all_results[skip : skip + limit]

    elif metric == "receivables":
        debt_q = select(Invoice).options(selectinload(Invoice.reservation)).where(Invoice.status.in_([InvoiceStatus.UNPAID, InvoiceStatus.PARTIAL, InvoiceStatus.APPROVED]))
        if start_date and end_date: debt_q = debt_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
        if rep_ids or region_id or product_id:
            debt_q = debt_q.join(Reservation, Invoice.reservation_id == Reservation.id)
            debt_q = apply_filters(debt_q, Reservation)
        debt_q = debt_q.where(func.coalesce(Invoice.total_amount, 0) > func.coalesce(Invoice.paid_amount, 0))
        rows = (await db.execute(debt_q.order_by(Invoice.date.desc()).offset(skip).limit(limit))).scalars().all()
        return [{"id": r.id, "date": r.date.isoformat() if r.date else "-", "invoice_num": r.factura_number, "total_amount": r.total_amount or 0, "paid_amount": r.paid_amount or 0, "debt_amount": (r.total_amount or 0) - (r.paid_amount or 0), "customer": r.reservation.customer_name if r.reservation else "-"} for r in rows]

    elif metric == "expenses":
        expense_q = select(OtherExpense).options(selectinload(OtherExpense.category), selectinload(OtherExpense.created_by))
        if start_date and end_date: expense_q = expense_q.where(and_(OtherExpense.date >= start_date, OtherExpense.date < end_date))
        if rep_ids: expense_q = expense_q.where(OtherExpense.created_by_id.in_(rep_ids))
        if region_id: expense_q = expense_q.where(OtherExpense.region_id == region_id)
        rows = (await db.execute(expense_q.order_by(OtherExpense.date.desc()).offset(skip).limit(limit))).scalars().all()
        return [{"id": r.id, "date": r.date.isoformat() if r.date else "-", "amount": r.amount, "category": r.category.name if r.category else "-", "description": r.comment or "-", "author": r.created_by.full_name if r.created_by else "-"} for r in rows]

    elif metric in ["bonus_accrued", "bonus_paid", "preinvest", "salary_accrued", "salary_paid", "salary_balance"]:
        # Determine which ledger_category to query
        ledger_cat = "salary" if metric.startswith("salary") else "bonus"

        from sqlalchemy.orm import selectinload as sil
        bonus_q = select(BonusLedger).options(
            sil(BonusLedger.doctor),
            sil(BonusLedger.user),
            sil(BonusLedger.product),
            sil(BonusLedger.payment).selectinload(Payment.invoice).selectinload(Invoice.reservation)
        ).join(User, BonusLedger.user_id == User.id).where(User.is_active == True, User.role == UserRole.MED_REP, BonusLedger.ledger_category == ledger_cat)
        
        if metric in ["bonus_accrued", "salary_accrued"]:
            bonus_q = bonus_q.where(and_(BonusLedger.ledger_type == LedgerType.ACCRUAL, or_(BonusLedger.notes != "Аванс (Предынвест)", BonusLedger.notes.is_(None))))
        elif metric in ["bonus_paid", "salary_paid"]:
            bonus_q = bonus_q.where(
                or_(
                    and_(BonusLedger.ledger_type == LedgerType.ACCRUAL, BonusLedger.is_paid == True),
                    BonusLedger.ledger_type == LedgerType.ADVANCE,
                    BonusLedger.ledger_type == LedgerType.PAYOUT
                )
            )
        elif metric in ["salary_balance"]:
            # Unpaid accruals = the remaining salary owed to the MedRep
            bonus_q = bonus_q.where(and_(BonusLedger.ledger_type == LedgerType.ACCRUAL, BonusLedger.is_paid == False))
        elif metric == "preinvest":
            bonus_q = bonus_q.where(and_(BonusLedger.ledger_type == LedgerType.ACCRUAL, BonusLedger.notes == "Аванс (Предынвест)"))
        if month and year:
            bonus_q = bonus_q.where(
                or_(
                    and_(BonusLedger.target_month == month, BonusLedger.target_year == year),
                    and_(BonusLedger.target_month.is_(None), extract('month', BonusLedger.created_at) == month, extract('year', BonusLedger.created_at) == year)
                )
            )
        elif quarter and year:
            start_m = (quarter - 1) * 3 + 1
            bonus_q = bonus_q.where(
                or_(
                    and_(BonusLedger.target_month >= start_m, BonusLedger.target_month <= start_m + 2, BonusLedger.target_year == year),
                    and_(BonusLedger.target_month.is_(None), extract('month', BonusLedger.created_at) >= start_m, extract('month', BonusLedger.created_at) <= start_m + 2, extract('year', BonusLedger.created_at) == year)
                )
            )
        elif year:
            bonus_q = bonus_q.where(
                or_(
                    BonusLedger.target_year == year,
                    and_(BonusLedger.target_year.is_(None), extract('year', BonusLedger.created_at) == year)
                )
            )
        elif start_date and end_date:
            bonus_q = bonus_q.where(and_(BonusLedger.created_at >= start_date, BonusLedger.created_at < end_date))
            
        if rep_ids: bonus_q = bonus_q.where(BonusLedger.user_id.in_(rep_ids))
        if region_id: bonus_q = bonus_q.join(Doctor, BonusLedger.doctor_id == Doctor.id).where(Doctor.region_id == region_id)
        if product_id: bonus_q = bonus_q.where(BonusLedger.product_id == product_id)
        rows = (await db.execute(bonus_q.order_by(BonusLedger.created_at.desc()).offset(skip).limit(limit))).scalars().all()

        result = []
        for r in rows:
            # Try to get payment/invoice info if bonus was triggered by a payment
            payment_info = None
            invoice_info = None
            if r.payment:
                p = r.payment
                payment_info = {
                    "payment_id": p.id,
                    "payment_amount": p.amount,
                    "payment_date": p.date.isoformat() if p.date else None,
                }
                if p.invoice:
                    inv = p.invoice
                    reservation = inv.reservation if inv else None
                    invoice_info = {
                        "invoice_id": inv.id,
                        "factura_number": inv.factura_number or f"#{inv.id}",
                        "invoice_total": inv.total_amount,
                        "invoice_paid": inv.paid_amount,
                        "customer": reservation.customer_name if reservation else "-",
                    }

            result.append({
                "id": r.id,
                "date": r.created_at.isoformat() if r.created_at else "-",
                "amount": r.amount,
                "doctor": r.doctor.full_name if r.doctor else "-",
                "med_rep": r.user.full_name if r.user else "-",
                "product": r.product.name if r.product else "-",
                "description": r.notes or "-",
                "payment": payment_info,
                "invoice": invoice_info,
            })
        return result

    elif metric == "gross_profit":
        gross_q = select(ReservationItem).options(selectinload(ReservationItem.product), selectinload(ReservationItem.reservation).selectinload(Reservation.invoice)).join(Reservation, ReservationItem.reservation_id == Reservation.id).join(Invoice, Invoice.reservation_id == Reservation.id).join(Product, ReservationItem.product_id == Product.id).where(and_(Invoice.total_amount > 0, Invoice.status != InvoiceStatus.CANCELLED))
        if start_date and end_date: gross_q = gross_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
        if rep_ids or region_id: gross_q = gross_q.outerjoin(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id)
        if rep_ids: gross_q = gross_q.where(or_(Reservation.created_by_id.in_(rep_ids), MedicalOrganization.assigned_reps.any(User.id.in_(rep_ids))))
        if region_id: gross_q = gross_q.where(MedicalOrganization.region_id == region_id)
        if product_id: gross_q = gross_q.where(ReservationItem.product_id == product_id)
        
        rows = (await db.execute(gross_q.order_by(Invoice.date.desc()).offset(skip).limit(limit))).scalars().all()
        res_payload = []
        for r in rows:
            paid_ratio = (r.reservation.invoice.paid_amount or 0) / r.reservation.invoice.total_amount if r.reservation and r.reservation.invoice and r.reservation.invoice.total_amount > 0 else 0
            sale_price = r.price * (1 - (r.discount_percent or 0) / 100.0)
            ratio = (r.price / r.product.price) if r.product and r.product.price and r.product.price > 0 else 1.0
            prod_price = (r.product.production_price or 0) * ratio
            if r.reservation and not r.reservation.is_salary_enabled:
                salary = 0
            else:
                salary = r.salary_amount if (r.salary_amount or 0) > 0 else (r.product.salary_expense or 0)
            
            if r.reservation and not r.reservation.is_bonus_eligible:
                marketing = 0
            else:
                marketing = r.marketing_amount if (r.marketing_amount or 0) > 0 else (r.product.marketing_expense or 0)
            other_per_unit = (r.product.other_expenses or 0) * ratio
            unit_profit = sale_price - prod_price - salary - marketing - other_per_unit
            total_profit_realized = (unit_profit * r.quantity) * paid_ratio
            if total_profit_realized > 0:
                res_payload.append({
                    "id": r.id,
                    "invoice_num": r.reservation.invoice.factura_number if r.reservation and r.reservation.invoice else "-",
                    "date": r.reservation.invoice.date.isoformat() if r.reservation and r.reservation.invoice else "-",
                    "product": r.product.name if r.product else "-",
                    "qty": r.quantity,
                    "paid_ratio": round(paid_ratio * 100, 1),
                    "profit": float(total_profit_realized)
                })
        return res_payload

    elif metric == "net_profit":
        # Net profit = gross profit per invoice item - total other expenses
        # Show breakdown: gross profit by invoice item, then expenses as a footer row
        gross_q = select(ReservationItem).options(
            selectinload(ReservationItem.product),
            selectinload(ReservationItem.reservation).selectinload(Reservation.invoice).selectinload(Invoice.payments),
            selectinload(ReservationItem.reservation).selectinload(Reservation.med_org).selectinload(MedicalOrganization.region),
            selectinload(ReservationItem.reservation).selectinload(Reservation.created_by)
        ).join(Reservation, ReservationItem.reservation_id == Reservation.id)\
         .join(Invoice, Invoice.reservation_id == Reservation.id)\
         .join(Product, ReservationItem.product_id == Product.id)\
         .where(and_(Invoice.total_amount > 0, Invoice.status != InvoiceStatus.CANCELLED))
        if start_date and end_date: gross_q = gross_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
        if rep_ids or region_id: gross_q = gross_q.outerjoin(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id)
        if rep_ids: gross_q = gross_q.where(or_(Reservation.created_by_id.in_(rep_ids), MedicalOrganization.assigned_reps.any(User.id.in_(rep_ids))))
        if region_id: gross_q = gross_q.where(MedicalOrganization.region_id == region_id)
        if product_id: gross_q = gross_q.where(ReservationItem.product_id == product_id)

        rows = (await db.execute(gross_q.order_by(Invoice.date.desc()).offset(skip).limit(limit))).scalars().all()
        res_payload = []
        for r in rows:
            inv = r.reservation.invoice if r.reservation else None
            if not inv or not inv.total_amount: continue
            paid_ratio = (inv.paid_amount or 0) / inv.total_amount
            # Цена продажи — faktik faktura narxi (chegirma hisobga olingan)
            sale_price = r.price * (1 - (r.discount_percent or 0) / 100.0)
            # Себестоимость — mahsulotning o'zgarmas tannarxi (ratio bilan ko'paytirish noto'g'ri!)
            prod_price = (r.product.production_price or 0)
            # Зарплата МП — fakturadagi yozuv, yo'q bo'lsa mahsulot default qiymati
            if r.reservation and not r.reservation.is_salary_enabled:
                salary = 0
            else:
                salary = r.salary_amount if (r.salary_amount or 0) > 0 else (r.product.salary_expense or 0)
            
            # Маркетинг — fakturadagi yozuv, yo'q bo'lsa mahsulot default qiymati
            if r.reservation and not r.reservation.is_bonus_eligible:
                marketing = 0
            else:
                marketing = r.marketing_amount if (r.marketing_amount or 0) > 0 else (r.product.marketing_expense or 0)
            # Boshqa xarajatlar — o'zgarmas
            other_per_unit = (r.product.other_expenses or 0)
            unit_profit = sale_price - prod_price - salary - marketing - other_per_unit
            total_profit_realized = (unit_profit * r.quantity) * paid_ratio
            region_name = (r.reservation.med_org.region.name if r.reservation and r.reservation.med_org and r.reservation.med_org.region else "-") if r.reservation else "-"
            med_rep_name = (r.reservation.created_by.full_name if r.reservation and r.reservation.created_by else "-")
            res_payload.append({
                "id": r.id,
                "invoice_num": inv.factura_number if inv else "-",
                "date": inv.date.isoformat() if inv else "-",
                "product": r.product.name if r.product else "-",
                "region": region_name,
                "med_rep": med_rep_name,
                "qty": r.quantity,
                "paid_ratio": round(paid_ratio * 100, 1),
                "gross_profit": round(float((unit_profit * r.quantity) * paid_ratio), 2),
                # For context: show formula breakdown
                "sale_price": round(float(sale_price), 2),
                "prod_price": round(float(prod_price), 2),
                "salary": round(float(salary), 2),
                "marketing": round(float(marketing), 2),
            })
        return res_payload

    elif metric == "sold_items":
        items_sold_q = select(ReservationItem).options(
            selectinload(ReservationItem.product),
            selectinload(ReservationItem.reservation).selectinload(Reservation.invoice),
            selectinload(ReservationItem.reservation).selectinload(Reservation.med_org).selectinload(MedicalOrganization.region),
            selectinload(ReservationItem.reservation).selectinload(Reservation.med_org).selectinload(MedicalOrganization.assigned_reps),
            selectinload(ReservationItem.reservation).selectinload(Reservation.created_by)
        ).join(Reservation, ReservationItem.reservation_id == Reservation.id)\
         .join(Invoice, Invoice.reservation_id == Reservation.id)\
         .where(Invoice.status != InvoiceStatus.CANCELLED)
         
        items_sold_q = apply_filters(items_sold_q, Reservation)
        if start_date and end_date: 
            items_sold_q = items_sold_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
        if product_id: 
            items_sold_q = items_sold_q.where(ReservationItem.product_id == int(product_id))
            
        rows = (await db.execute(items_sold_q.order_by(Invoice.date.desc()).offset(skip).limit(limit))).scalars().all()
        
        res_payload = []
        for r in rows:
            inv = r.reservation.invoice if r.reservation else None
            region_name = (r.reservation.med_org.region.name if r.reservation and r.reservation.med_org and r.reservation.med_org.region else "-") if r.reservation else "-"
            
            med_org = r.reservation.med_org if r.reservation else None
            if med_org and med_org.assigned_reps:
                med_rep_name = ", ".join([u.full_name for u in med_org.assigned_reps if u.full_name])
            elif r.reservation and r.reservation.created_by:
                med_rep_name = r.reservation.created_by.full_name
            else:
                med_rep_name = "-"

            customer = (inv.reservation.med_org.name if inv.reservation and inv.reservation.med_org else (inv.reservation.customer_name if inv.reservation else "-")) if inv else "-"
            
            nds_percent = (r.reservation.nds_percent if r.reservation and r.reservation.nds_percent is not None else 12.0)
            nds_multiplier = 1 + (nds_percent / 100.0)
            
            price_with_nds = round((r.price or 0) * nds_multiplier, 2)
            total_with_nds = round(price_with_nds * (r.quantity or 0), 2)
            
            res_payload.append({
                "id": r.id,
                "date": inv.date.isoformat() if inv and inv.date else "-",
                "customer": customer,
                "invoice_num": inv.factura_number if inv else "-",
                "product": r.product.name if r.product else "-",
                "region": region_name,
                "med_rep": med_rep_name,
                "qty": r.quantity,
                "sale_price": price_with_nds,
                "total_amount": total_with_nds
            })
        return res_payload

    return []

@router.get("/dashboard/director-report-excel")
async def get_director_report_excel(
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
    month: int = None,
    year: int = None
) -> Any:
    """
    Director Excel report: Product Manager → Regional Manager → MedRep
    Columns: per-product Plan(qty), Fact(qty), Fact(sum) + totals
    """
    import io, calendar as _cal
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    from app.models.product import Product
    from app.models.sales import Plan, Invoice, Reservation, ReservationItem
    from sqlalchemy.orm import selectinload
    from sqlalchemy import select, and_, extract, func

    if current_user.role not in [UserRole.INVESTOR, UserRole.DIRECTOR, UserRole.DEPUTY_DIRECTOR, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    if not month:
        month = datetime.utcnow().month
    if not year:
        year = datetime.utcnow().year

    # ── 1. Products (columns) ───────────────────────────────────────────────────
    products = (await db.execute(select(Product).where(Product.is_active == True).order_by(Product.name))).scalars().all()
    prod_count = len(products)

    # ── 2. MedReps with full manager chain ─────────────────────────────────────
    medreps_res = await db.execute(
        select(User)
        .options(
            selectinload(User.manager).selectinload(User.manager)
        )
        .where(User.role == UserRole.MED_REP, User.is_active == True)
        .order_by(User.full_name)
    )
    medreps = medreps_res.scalars().all()

    # Build hierarchy map: pm_name -> rm_name -> [medrep]
    from collections import defaultdict
    hierarchy: dict[str, dict[str, list]] = defaultdict(lambda: defaultdict(list))
    for mr in medreps:
        rm_name = "-"
        pm_name = "-"
        mgr = mr.manager
        if mgr:
            rm_name = mgr.full_name or mgr.username
            mgr2 = mgr.manager
            if mgr2 and mgr2.role == UserRole.PRODUCT_MANAGER:
                pm_name = mgr2.full_name or mgr2.username
        hierarchy[pm_name][rm_name].append(mr)

    # ── 3. Plans (month/year, by med_rep_id + product_id) ──────────────────────
    plan_q = select(Plan).where(and_(Plan.month == month, Plan.year == year))
    plans = (await db.execute(plan_q)).scalars().all()
    # plan_map[(med_rep_id, product_id)] = (qty, amount)
    plan_map: dict[tuple, tuple] = {}
    for p in plans:
        if p.med_rep_id:
            key = (p.med_rep_id, p.product_id)
            old_q, old_a = plan_map.get(key, (0, 0))
            plan_map[key] = (old_q + (p.target_quantity or 0), old_a + (p.target_amount or 0))

    # ── 4. Facts: Invoice+ReservationItem by med_rep + product for this period ──
    first_day = datetime(year, month, 1)
    last_day  = datetime(year, month, _cal.monthrange(year, month)[1], 23, 59, 59)

    fact_q = (
        select(
            Reservation.created_by_id.label("med_rep_id"),
            ReservationItem.product_id,
            func.sum(ReservationItem.quantity).label("qty"),
            func.sum(ReservationItem.total_price).label("total_sum"),
        )
        .join(ReservationItem, ReservationItem.reservation_id == Reservation.id)
        .join(Invoice, Invoice.reservation_id == Reservation.id)
        .where(
            Invoice.date.between(first_day, last_day),
            Invoice.status.notin_(["cancelled", "returned"])
        )
        .group_by(Reservation.created_by_id, ReservationItem.product_id)
    )
    fact_rows = (await db.execute(fact_q)).all()
    # fact_map[(med_rep_id, product_id)] = (qty, sum)
    fact_map: dict[tuple, tuple] = {}
    for row in fact_rows:
        if row.med_rep_id:
            fact_map[(row.med_rep_id, row.product_id)] = (int(row.qty or 0), float(row.total_sum or 0))

    # ── 5. Build Excel ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Отчет {month}-{year}"

    # Styles
    C_BLUE   = "1E3A5F";  C_BLUE_L  = "BDD7EE"
    C_GREEN  = "375623";  C_GREEN_L = "E2EFDA"
    C_GREY   = "404040";  C_GREY_L  = "F2F2F2"
    C_YELLOW = "7F6000";  C_YELLOW_L= "FFEB9C"
    C_WHITE  = "FFFFFF"

    def fill(hex_color):
        return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    def font(bold=False, color=C_WHITE, size=9):
        return Font(bold=bold, color=color, size=size)

    thin  = Side(style="thin",   color="999999")
    thick = Side(style="medium", color="555555")
    def border(left=thin, right=thin, top=thin, bottom=thin):
        return Border(left=left, right=right, top=top, bottom=bottom)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")
    right_a= Alignment(horizontal="right",  vertical="center")
    vert   = Alignment(text_rotation=90, horizontal="center", vertical="center")

    # ── Header row 1: static cols + per-product groups ─────────────────────────
    STATIC_COLS = ["Продукт Менеджер", "Рег. Менеджер / МП", "Регион"]
    S = len(STATIC_COLS)  # number of static columns = 3

    # Row 1: static headers (merge rows 1-3) + product group headers (merge cols per product)
    # Row 2: per-product sub-headers: Plan / Факт (дон.) / Факт (сум)
    # Row 3: empty (continuation of merge)

    # Static header cells (merge 3 rows)
    for ci, h in enumerate(STATIC_COLS, 1):
        ws.merge_cells(start_row=1, start_column=ci, end_row=3, end_column=ci)
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = font(bold=True, color=C_WHITE)
        cell.fill      = fill(C_BLUE)
        cell.alignment = vert
        cell.border    = border()

    # Per-product group headers (3 cols each: plan_qty, fact_qty, fact_sum)
    COLS_PER_PROD = 3
    for pi, prod in enumerate(products):
        start_col = S + 1 + pi * COLS_PER_PROD
        end_col   = start_col + COLS_PER_PROD - 1

        # Row 1: product name (merge 3 cols)
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        c = ws.cell(row=1, column=start_col, value=prod.name)
        c.font = font(bold=True, color=C_WHITE); c.fill = fill(C_GREEN); c.alignment = center; c.border = border()

        # Row 2: sub-labels
        sub_labels = ["План (дон.)", "Факт (дон.)", "Факт (сум)"]
        for si, lbl in enumerate(sub_labels):
            c2 = ws.cell(row=2, column=start_col + si, value=lbl)
            c2.font = font(bold=True, color=C_GREEN); c2.fill = fill(C_GREEN_L)
            c2.alignment = center; c2.border = border()

        # Row 3: % выполнения (colspan 3)
        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
        c3 = ws.cell(row=3, column=start_col, value="% выполнения")
        c3.font = font(bold=False, color=C_GREEN); c3.fill = fill(C_GREEN_L)
        c3.alignment = center; c3.border = border()

    # Totals columns (after all products): total plan qty, total fact qty, total fact sum
    tc_start = S + 1 + prod_count * COLS_PER_PROD
    ws.merge_cells(start_row=1, start_column=tc_start, end_row=1, end_column=tc_start + 2)
    c = ws.cell(row=1, column=tc_start, value="ИТОГО")
    c.font = font(bold=True); c.fill = fill(C_GREY); c.alignment = center; c.border = border()
    for si, lbl in enumerate(["План (дон.)", "Факт (дон.)", "Факт (сум)"]):
        c2 = ws.cell(row=2, column=tc_start + si, value=lbl)
        c2.font = font(bold=True, color=C_GREY); c2.fill = fill(C_GREY_L); c2.alignment = center; c2.border = border()
    ws.merge_cells(start_row=3, start_column=tc_start, end_row=3, end_column=tc_start + 2)
    c3 = ws.cell(row=3, column=tc_start, value="% выполнения")
    c3.font = font(bold=False, color=C_GREY); c3.fill = fill(C_GREY_L); c3.alignment = center; c3.border = border()

    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18

    # ── Data rows ───────────────────────────────────────────────────────────────
    def pct(plan_q, fact_q):
        if not plan_q: return "—"
        return f"{round(fact_q / plan_q * 100)}%"

    def write_data_row(row_idx, pm, rm_mr_label, region_label,
                       row_fill_color, row_font_color, bold=False):
        """Write one data row; returns (total_plan_q, total_fact_q, total_fact_s)."""
        for ci, val in enumerate([pm, rm_mr_label, region_label], 1):
            c = ws.cell(row=row_idx, column=ci, value=val)
            c.fill = fill(row_fill_color); c.font = font(bold=bold, color=row_font_color)
            c.alignment = left; c.border = border()
        return row_idx  # caller fills product columns

    current_row = 4

    for pm_name in sorted(hierarchy.keys()):
        rm_dict = hierarchy[pm_name]

        # Accumulators for PM summary row
        pm_plan_tot = 0; pm_fact_qty_tot = 0; pm_fact_sum_tot = 0
        pm_per_prod_plan = [0]*prod_count
        pm_per_prod_fqty = [0]*prod_count
        pm_per_prod_fsum = [0.0]*prod_count

        pm_row_idx = current_row  # we'll fill it after writing RM/MR rows
        current_row += 1  # reserve PM row

        for rm_name in sorted(rm_dict.keys()):
            mr_list = rm_dict[rm_name]

            rm_plan_tot = 0; rm_fact_qty_tot = 0; rm_fact_sum_tot = 0
            rm_per_prod_plan = [0]*prod_count
            rm_per_prod_fqty = [0]*prod_count
            rm_per_prod_fsum = [0.0]*prod_count

            rm_row_idx = current_row
            current_row += 1  # reserve RM row

            for mr in mr_list:
                mr_row = current_row
                current_row += 1

                # Static cells
                for ci, val in enumerate(["", f"    {mr.full_name or mr.username}", ""], 1):
                    c = ws.cell(row=mr_row, column=ci, value=val)
                    c.fill = fill(C_WHITE); c.font = font(bold=False, color="333333")
                    c.alignment = left; c.border = border()

                mr_plan_tot = 0; mr_fact_qty_tot = 0; mr_fact_sum_tot = 0

                for pi, prod in enumerate(products):
                    plan_q, _ = plan_map.get((mr.id, prod.id), (0, 0))
                    fact_q, fact_s = fact_map.get((mr.id, prod.id), (0, 0.0))
                    col_base = S + 1 + pi * COLS_PER_PROD

                    for si, (val, fmt) in enumerate([(plan_q, None), (fact_q, None), (fact_s, '#,##0')]):
                        c = ws.cell(row=mr_row, column=col_base + si, value=val)
                        c.fill = fill(C_WHITE); c.font = font(bold=False, color="333333")
                        c.alignment = right_a; c.border = border()
                        if fmt: c.number_format = fmt

                    # % выполнения
                    p_col = S + 1 + pi * COLS_PER_PROD  # col_base start
                    # spans 3 cols but we write to first col; we skip merging for perf
                    # write pct in a merged sense — put value only in first of 3 cols logically
                    # Actually we'll skip a separate pct row per medrep, include in sub-label

                    mr_plan_tot      += plan_q
                    mr_fact_qty_tot  += fact_q
                    mr_fact_sum_tot  += fact_s
                    rm_per_prod_plan[pi] += plan_q
                    rm_per_prod_fqty[pi] += fact_q
                    rm_per_prod_fsum[pi] += fact_s

                # MedRep totals
                tc = S + 1 + prod_count * COLS_PER_PROD
                for si, val in enumerate([mr_plan_tot, mr_fact_qty_tot, mr_fact_sum_tot]):
                    c = ws.cell(row=mr_row, column=tc + si, value=val)
                    c.fill = fill(C_WHITE); c.font = font(bold=False, color="333333")
                    c.alignment = right_a; c.border = border()
                    if si == 2: c.number_format = '#,##0'

                rm_plan_tot     += mr_plan_tot
                rm_fact_qty_tot += mr_fact_qty_tot
                rm_fact_sum_tot += mr_fact_sum_tot

            # ── RM summary row ──────────────────────────────────────────────
            for ci, val in enumerate(["", f"  RM: {rm_name}", ""], 1):
                c = ws.cell(row=rm_row_idx, column=ci, value=val)
                c.fill = fill(C_BLUE_L); c.font = font(bold=True, color=C_BLUE)
                c.alignment = left; c.border = border(bottom=thick)

            for pi in range(prod_count):
                col_base = S + 1 + pi * COLS_PER_PROD
                for si, val in enumerate([rm_per_prod_plan[pi], rm_per_prod_fqty[pi], rm_per_prod_fsum[pi]]):
                    c = ws.cell(row=rm_row_idx, column=col_base + si, value=val)
                    c.fill = fill(C_BLUE_L); c.font = font(bold=True, color=C_BLUE)
                    c.alignment = right_a; c.border = border(bottom=thick)
                    if si == 2: c.number_format = '#,##0'

            tc = S + 1 + prod_count * COLS_PER_PROD
            for si, val in enumerate([rm_plan_tot, rm_fact_qty_tot, rm_fact_sum_tot]):
                c = ws.cell(row=rm_row_idx, column=tc + si, value=val)
                c.fill = fill(C_BLUE_L); c.font = font(bold=True, color=C_BLUE)
                c.alignment = right_a; c.border = border(bottom=thick)
                if si == 2: c.number_format = '#,##0'

            pm_plan_tot     += rm_plan_tot
            pm_fact_qty_tot += rm_fact_qty_tot
            pm_fact_sum_tot += rm_fact_sum_tot
            for pi in range(prod_count):
                pm_per_prod_plan[pi] += rm_per_prod_plan[pi]
                pm_per_prod_fqty[pi] += rm_per_prod_fqty[pi]
                pm_per_prod_fsum[pi] += rm_per_prod_fsum[pi]

        # ── PM summary row ──────────────────────────────────────────────────
        for ci, val in enumerate([f"PM: {pm_name}", "", ""], 1):
            c = ws.cell(row=pm_row_idx, column=ci, value=val)
            c.fill = fill(C_BLUE); c.font = font(bold=True, color=C_WHITE)
            c.alignment = left; c.border = border(top=thick, bottom=thick)

        for pi in range(prod_count):
            col_base = S + 1 + pi * COLS_PER_PROD
            pq = pm_per_prod_plan[pi]; fq = pm_per_prod_fqty[pi]; fs = pm_per_prod_fsum[pi]
            for si, val in enumerate([pq, fq, fs]):
                c = ws.cell(row=pm_row_idx, column=col_base + si, value=val)
                c.fill = fill(C_BLUE); c.font = font(bold=True, color=C_WHITE)
                c.alignment = right_a; c.border = border(top=thick, bottom=thick)
                if si == 2: c.number_format = '#,##0'

        tc = S + 1 + prod_count * COLS_PER_PROD
        for si, val in enumerate([pm_plan_tot, pm_fact_qty_tot, pm_fact_sum_tot]):
            c = ws.cell(row=pm_row_idx, column=tc + si, value=val)
            c.fill = fill(C_BLUE); c.font = font(bold=True, color=C_WHITE)
            c.alignment = right_a; c.border = border(top=thick, bottom=thick)
            if si == 2: c.number_format = '#,##0'

    # ── Grand total row ─────────────────────────────────────────────────────────
    grand_plan = sum(plan_map[k][0] for k in plan_map)
    grand_fqty = sum(v[0] for v in fact_map.values())
    grand_fsum = sum(v[1] for v in fact_map.values())

    gt_row = current_row
    for ci, val in enumerate(["ЖAMI / ИТОГО", "", ""], 1):
        c = ws.cell(row=gt_row, column=ci, value=val)
        c.fill = fill(C_YELLOW); c.font = font(bold=True, color=C_YELLOW)
        c.alignment = left; c.border = border(top=thick)

    for pi, prod in enumerate(products):
        col_base = S + 1 + pi * COLS_PER_PROD
        pq = sum(plan_map.get((mr.id, prod.id), (0,0))[0] for pm in hierarchy for rm in hierarchy[pm] for mr in hierarchy[pm][rm])
        fq = sum(fact_map.get((mr.id, prod.id), (0,0))[0] for pm in hierarchy for rm in hierarchy[pm] for mr in hierarchy[pm][rm])
        fs = sum(fact_map.get((mr.id, prod.id), (0,0.0))[1] for pm in hierarchy for rm in hierarchy[pm] for mr in hierarchy[pm][rm])
        for si, val in enumerate([pq, fq, fs]):
            c = ws.cell(row=gt_row, column=col_base + si, value=val)
            c.fill = fill(C_YELLOW_L); c.font = font(bold=True, color=C_YELLOW)
            c.alignment = right_a; c.border = border(top=thick)
            if si == 2: c.number_format = '#,##0'

    tc = S + 1 + prod_count * COLS_PER_PROD
    for si, val in enumerate([grand_plan, grand_fqty, grand_fsum]):
        c = ws.cell(row=gt_row, column=tc + si, value=val)
        c.fill = fill(C_YELLOW_L); c.font = font(bold=True, color=C_YELLOW)
        c.alignment = right_a; c.border = border(top=thick)
        if si == 2: c.number_format = '#,##0'

    # ── Column widths ───────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 22  # PM
    ws.column_dimensions[get_column_letter(2)].width = 26  # RM/MR name
    ws.column_dimensions[get_column_letter(3)].width = 16  # Region
    total_cols = S + prod_count * COLS_PER_PROD + 3
    for ci in range(S + 1, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    ws.freeze_panes = "A4"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Director_Report_{year}_{month:02d}.xlsx"
    resp_headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=resp_headers
    )

