from typing import Any, List, Optional, Dict
import io
import logging
import urllib.parse

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.api import deps
from app.crud import crud_sales
from datetime import datetime
from app.models.sales import Reservation, ReservationItem, Invoice, InvoiceStatus, ReservationStatus, Payment
from app.models.user import User, UserRole
from app.models.crm import MedicalOrganization, Region
from app.models.product import Product
from app.models.warehouse import Warehouse
from app.schemas.sales import (
    Plan, PlanCreate, 
    Reservation as ReservationSchema, ReservationCreate, ReservationUpdate,
    Invoice as InvoiceSchema, Payment as PaymentSchema, PaymentCreate,
    DoctorFactAssignment as DoctorFactAssignmentSchema, DoctorFactAssignmentCreate, SaleFact,
    BonusPayment as BonusPaymentSchema, BonusPaymentCreate, BonusPaymentUpdate,
    ReservationReturnCreate, BonusAllocationCreate
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import traceback

router = APIRouter()

# Plans
@router.post("/plans/", response_model=Plan)
async def create_plan(
    *,
    db: AsyncSession = Depends(deps.get_db),
    plan_in: PlanCreate,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    plan = await crud_sales.create_plan(db, obj_in=plan_in)
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "CREATE", "Plan", plan.id,
        f"План создан: {plan.target_quantity} шт., Месяц: {plan.month}/{plan.year}",
        request
    )
    return plan

@router.get("/plans/", response_model=List[Plan])
async def read_plans(
    db: AsyncSession = Depends(deps.get_db),
    skip: int = 0,
    limit: int = 100,
    month: int = None,
    year: int = None,
    med_rep_id: int = None,
    doctor_id: int = None,
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    if current_user.role == UserRole.MED_REP:
        med_rep_id = current_user.id
    
    return await crud_sales.get_plans(db, skip=skip, limit=limit, month=month, year=year, med_rep_id=med_rep_id, doctor_id=doctor_id)

@router.put("/plans/{id}", response_model=Plan)
async def update_plan(
    *,
    db: AsyncSession = Depends(deps.get_db),
    id: int,
    plan_in: dict,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    plan = await crud_sales.get_plan(db, id=id)
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
    
    updated_plan = await crud_sales.update_plan(db, db_obj=plan, obj_in=plan_in)
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "UPDATE", "Plan", updated_plan.id,
        f"План изменен: ID {id}",
        request
    )
    return updated_plan

@router.delete("/plans/{id}")
async def delete_plan(
    *,
    db: AsyncSession = Depends(deps.get_db),
    id: int,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    plan = await crud_sales.get_plan(db, id=id)
    if not plan:
        raise HTTPException(status_code=404, detail="Plan not found")
        
    await crud_sales.delete_plan(db, id=id)
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "DELETE", "Plan", id,
        f"План удален: ID {id}",
        request
    )
    return {"ok": True}


# Reservations (Bron)
@router.post("/reservations/", response_model=ReservationSchema)
async def create_reservation(
    *,
    db: AsyncSession = Depends(deps.get_db),
    reservation_in: ReservationCreate,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    # Use the service that locks and deducts stock
    from app.services.reservation_service import ReservationService
    reservation, mod_summary = await ReservationService.create_reservation_with_stock_lock(
        db=db, 
        obj_in=reservation_in, 
        user_id=current_user.id
    )
    
    log_description = f"Бронь создана (Склад зарезервирован): ID {reservation.id}, Сумма: {getattr(reservation, 'total_amount', 0) or 0:,.0f} UZS"
    if mod_summary:
        log_description += f" | Изменения: {mod_summary}"

    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "CREATE", "Reservation", reservation.id,
        log_description,
        request
    )
    return reservation

@router.delete("/reservations/{id}")
async def delete_reservation(
    *,
    db: AsyncSession = Depends(deps.get_db),
    id: int,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    if current_user.role not in [UserRole.INVESTOR, UserRole.DEPUTY_DIRECTOR, UserRole.HEAD_OF_ORDERS, UserRole.MED_REP, UserRole.DIRECTOR, UserRole.ADMIN]:
        raise HTTPException(status_code=403, detail="Not enough permissions to delete reservations.")
    
    # Check roles that require approval for deletion
    if current_user.role in [UserRole.HEAD_OF_ORDERS, UserRole.MED_REP, UserRole.DEPUTY_DIRECTOR, UserRole.FIELD_FORCE_MANAGER, UserRole.PRODUCT_MANAGER, UserRole.REGIONAL_MANAGER]:
        from app.models.sales import Reservation, Invoice
        res_query = select(Reservation).options(selectinload(Reservation.invoice)).where(Reservation.id == id)
        res_exc = await db.execute(res_query)
        reservation = res_exc.scalar_one_or_none()
        
        if not reservation:
            raise HTTPException(status_code=404, detail="Reservation not found")
        
        # If reservation has an invoice, mark the invoice for deletion instead
        if reservation.invoice:
            if (reservation.invoice.paid_amount or 0) > 0 or reservation.invoice.status == InvoiceStatus.PAID:
                raise HTTPException(status_code=400, detail="Нельзя удалить оплаченную счет-фактуру. Сначала отмените платежи.")
            
            reservation.invoice.is_deletion_pending = True
            reservation.invoice.deletion_requested_by_id = current_user.id
        else:
            reservation.is_deletion_pending = True
            reservation.deletion_requested_by_id = current_user.id
        
        await db.commit()
        
        from app.services.audit_service import log_action
        await log_action(
            db, current_user, "DELETE_REQUESTED", "Reservation", id,
            f"Запрошено удаление брони #{id}. Ожидает подтверждения склада.",
            request
        )
        return {"ok": True, "message": "Запрос на удаление отправлен заведующему склада (Deletion request sent to Warehouse Head)."}

    # Roles that can delete immediately: DIRECTOR, ADMIN, HEAD_OF_WAREHOUSE
    from app.services.reservation_service import ReservationService
    await ReservationService.cancel_reservation(db=db, reservation_id=id)
    
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "DELETE", "Reservation", id,
        f"Бронь отменена (Склад восстановлен): ID {id}",
        request
    )
    return {"ok": True, "message": "Reservation cancelled and stock restored."}

@router.get("/reservations/", response_model=List[ReservationSchema])
async def read_reservations(
    db: AsyncSession = Depends(deps.get_db),
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(deps.get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    med_rep_name: Optional[str] = None,
    med_org_name: Optional[str] = None,
    med_org_type: Optional[str] = None,
    is_tovar_skidka: Optional[bool] = None,
    inv_num: Optional[str] = None,
    status: Optional[str] = None,
    med_rep_id: Optional[int] = None,
    med_org_id: Optional[int] = None,
    region_id: Optional[int] = None
) -> Any:
    """
    Retrieve reservations with optional filtering.
    
    Roles Permitted and Filtering:
    - MED_REP: Sees only their own reservations or those assigned to their medical organizations.
    - PRODUCT_MANAGER/FF_MANAGER/REGIONAL_MANAGER: See reservations within their team hierarchy.
    - DIRECTOR/DEPUTY: See all reservations.
    
    Filters:
    - date_from, date_to: Range filter for reservation date.
    - med_rep_name: Search by MedRep's full name.
    - med_org_name: Search by Medical Organization name.
    - status: Filter by ReservationStatus (draft, pending, approved, etc.)
    """
    # Prioritize provided med_rep_id, but respect current_user roles
    final_med_rep_id = med_rep_id
    med_rep_ids = None
    if current_user.role == UserRole.MED_REP:
        final_med_rep_id = current_user.id
    elif current_user.role in [UserRole.PRODUCT_MANAGER, UserRole.FIELD_FORCE_MANAGER, UserRole.REGIONAL_MANAGER]:
        from app.crud import crud_user
        med_rep_ids = await crud_user.get_descendant_ids(db, current_user.id)
        if not med_rep_ids:
            med_rep_ids = [-1]
        final_med_rep_id = None
    
    # Regional Restriction for RM
    final_region_ids = [r.id for r in current_user.assigned_regions] if current_user.assigned_regions else None
    if current_user.role == UserRole.REGIONAL_MANAGER:
        if region_id:
            if region_id in (final_region_ids or []):
                final_region_ids = [region_id]
            else:
                final_region_ids = [-1] # No access
    elif region_id:
        final_region_ids = [region_id]

    dt_from = None
    if date_from and isinstance(date_from, str) and date_from.strip():
        try:
            dt_from = datetime.fromisoformat(date_from)
        except (ValueError, TypeError):
            dt_from = None
            
    dt_to = None
    if date_to and isinstance(date_to, str) and date_to.strip():
        try:
            dt_to = datetime.fromisoformat(date_to)
        except (ValueError, TypeError):
            dt_to = None
    
    return await crud_sales.get_reservations(
        db, 
        skip=skip, 
        limit=limit, 
        med_rep_id=final_med_rep_id,
        date_from=dt_from,
        date_to=dt_to,
        med_rep_name=med_rep_name,
        med_org_name=med_org_name,
        med_org_type=med_org_type,
        is_tovar_skidka=is_tovar_skidka,
        inv_num=inv_num,
        med_rep_ids=med_rep_ids,
        status=status,
        med_org_id=med_org_id,
        region_ids=final_region_ids
    )

@router.get("/reservations/{id}")
async def read_reservation(
    *,
    db: AsyncSession = Depends(deps.get_db),
    id: int,
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    try:
        query = select(Reservation).options(
            selectinload(Reservation.items).selectinload(ReservationItem.product).selectinload(Product.manufacturers),
            selectinload(Reservation.items).selectinload(ReservationItem.product).selectinload(Product.category),
            selectinload(Reservation.med_org).selectinload(MedicalOrganization.region),
            selectinload(Reservation.med_org).selectinload(MedicalOrganization.assigned_reps),
            selectinload(Reservation.created_by),
            selectinload(Reservation.warehouse).selectinload(Warehouse.stocks), selectinload(Reservation.warehouse).selectinload(Warehouse.med_org),
            selectinload(Reservation.invoice).selectinload(Invoice.payments).selectinload(Payment.processed_by)
        ).where(Reservation.id == id)
        
        result = await db.execute(query)
        reservation = result.scalar_one_or_none()
        
        if not reservation:
            raise HTTPException(status_code=404, detail="Reservation not found")
        
        # Manually serialize to avoid Pydantic issues with lazy loading
        inv = reservation.invoice
        resp = {
            "id": reservation.id,
            "customer_name": reservation.customer_name,
            "date": reservation.date.isoformat() if reservation.date else None,
            "total_amount": reservation.total_amount,
            "nds_percent": reservation.nds_percent,
            "status": reservation.status,
            "created_by": {"id": reservation.created_by.id, "full_name": reservation.created_by.full_name} if reservation.created_by else None,
            "med_org": {"id": reservation.med_org.id, "name": reservation.med_org.name, "inn": getattr(reservation.med_org, "inn", None)} if reservation.med_org else None,
            "items": [
                {
                    "id": item.id,
                    "quantity": item.quantity,
                    "price": item.price,
                    "marketing_amount": item.marketing_amount,
                    "default_marketing_amount": item.product.marketing_expense if item.product else 0,
                    "total_price": item.total_price,
                    "product": {"id": item.product.id, "name": item.product.name} if item.product else None
                }
                for item in (reservation.items or [])
            ],
            "invoice": {
                "id": inv.id,
                "total_amount": inv.total_amount,
                "paid_amount": inv.paid_amount,
                "status": inv.status,
                "realization_date": inv.realization_date.isoformat() if inv.realization_date else None,
                "payments": [
                    {"id": p.id, "amount": p.amount, "date": p.date.isoformat() if p.date else None}
                    for p in (inv.payments or [])
                ]
            } if inv else None
        }
        return resp
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())

@router.patch("/reservations/{id}/status", response_model=ReservationSchema)
async def update_reservation_status(
    *,
    db: AsyncSession = Depends(deps.get_db),
    id: int,
    status_update: ReservationUpdate,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    if current_user.role not in [UserRole.INVESTOR, UserRole.DEPUTY_DIRECTOR, UserRole.HEAD_OF_ORDERS]:
        raise HTTPException(status_code=400, detail="Not enough permissions")
    
    if status_update.status == ReservationStatus.APPROVED:
        from app.services.reservation_service import ReservationService
        updated_reservation = await ReservationService.activate_reservation(db, id)
    else:
        reservation = await crud_sales.get_reservation(db, id=id)
        if not reservation:
            raise HTTPException(status_code=404, detail="Reservation not found")
        updated_reservation = await crud_sales.update_reservation_status(db, db_obj=reservation, status=status_update.status)
    
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "UPDATE_STATUS", "Reservation", id,
        f"Статус брони изменен: {status_update.status}",
        request
    )
    return updated_reservation

# Invoices (Factura)
@router.post("/reservations/{id}/return", response_model=ReservationSchema)
async def map_return_reservation_items(
    *,
    db: AsyncSession = Depends(deps.get_db),
    id: int,
    return_in: ReservationReturnCreate,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    reservation = await crud_sales.request_return_reservation_items(
        db=db, 
        reservation_id=id, 
        obj_in=return_in, 
        user_id=current_user.id
    )
    from app.services.audit_service import log_action
    items_count = sum(r.quantity for r in return_in.items)
    await log_action(
        db, current_user, "RETURN_REQUESTED", "Reservation", id,
        f"Запрошен возврат по брони: {items_count} товаров. Ожидает одобрения склада.",
        request
    )
    return reservation

# Invoices (Factura)
@router.get("/invoices/", response_model=List[InvoiceSchema])
async def read_invoices(
    db: AsyncSession = Depends(deps.get_db),
    skip: int = 0,
    limit: int = 100,
    current_user: User = Depends(deps.get_current_user),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    med_rep_name: Optional[str] = None,
    med_org_name: Optional[str] = None,
    med_org_type: Optional[str] = None,
    is_tovar_skidka: Optional[bool] = None,
    inv_num: Optional[str] = None,
    status: Optional[str] = None,
    med_rep_id: Optional[int] = None,
    med_org_id: Optional[int] = None,
    has_debt: bool = False,
) -> Any:
    try:
        med_rep_ids = None
        if current_user.role == UserRole.MED_REP:
            med_rep_id = current_user.id
        elif current_user.role in [UserRole.PRODUCT_MANAGER, UserRole.FIELD_FORCE_MANAGER, UserRole.REGIONAL_MANAGER]:
            from app.crud import crud_user
            med_rep_ids = await crud_user.get_descendant_ids(db, current_user.id)
            if not med_rep_ids:
                med_rep_ids = [-1]
            med_rep_id = None

        region_ids = [r.id for r in current_user.assigned_regions] if current_user.assigned_regions else None
        
        # Robust date parsing
        dt_from = None
        if date_from and date_from.strip():
            try: dt_from = datetime.fromisoformat(date_from)
            except: pass
            
        dt_to = None
        if date_to and date_to.strip():
            try: dt_to = datetime.fromisoformat(date_to)
            except: pass
        
        return await crud_sales.get_invoices(
            db, 
            skip=skip, 
            limit=limit, 
            med_rep_id=med_rep_id,
            date_from=dt_from,
            date_to=dt_to,
            med_rep_name=med_rep_name,
            med_org_name=med_org_name,
            med_org_type=med_org_type,
            is_tovar_skidka=is_tovar_skidka,
            inv_num=inv_num,
            med_rep_ids=med_rep_ids,
            status=status,
            has_debt=has_debt,
            med_org_id=med_org_id,
            region_ids=region_ids
        )
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())

@router.get("/invoices/eligible-for-tovar-skidka", response_model=List[InvoiceSchema])
async def get_eligible_invoices_for_tovar_skidka(
    med_org_id: int,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    """
    Get invoices that are 100% paid and have unused promo balance.
    Filtered by medical organization.
    """
    try:
        query = select(Invoice).join(
            Reservation, Invoice.reservation_id == Reservation.id
        ).options(
            selectinload(Invoice.reservation).selectinload(Reservation.items).selectinload(ReservationItem.product).selectinload(Product.manufacturers),
            selectinload(Invoice.reservation).selectinload(Reservation.items).selectinload(ReservationItem.product).selectinload(Product.category),
            selectinload(Invoice.reservation).selectinload(Reservation.med_org).selectinload(MedicalOrganization.region),
            selectinload(Invoice.reservation).selectinload(Reservation.med_org).selectinload(MedicalOrganization.assigned_reps),
            selectinload(Invoice.reservation).selectinload(Reservation.warehouse).selectinload(Warehouse.stocks), selectinload(Reservation.warehouse).selectinload(Warehouse.med_org),
            selectinload(Invoice.reservation).selectinload(Reservation.created_by),
            selectinload(Invoice.payments).selectinload(Payment.processed_by)
        ).where(
            Reservation.med_org_id == med_org_id,
            Invoice.status == InvoiceStatus.PAID,
            Invoice.promo_balance > 0
        )
        result = await db.execute(query)
        return result.scalars().all()
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        logger.error(f"Error in get_eligible_invoices_for_tovar_skidka: {error_detail}")
        raise HTTPException(status_code=500, detail="Internal server error")

# Payments
@router.post("/payments/", response_model=PaymentSchema)
async def create_payment(
    *,
    db: AsyncSession = Depends(deps.get_db),
    payment_in: PaymentCreate,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    from app.services.finance_service import FinancialService
    payment = await FinancialService.process_payment(db, obj_in=payment_in, processor_id=current_user.id)
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "CREATE", "Payment", payment.id,
        f"Оплата принята: {payment.amount:,.0f} UZS",
        request
    )
    return payment

# Facts & Doctor Assignments
@router.get("/facts/", response_model=List[SaleFact])
async def read_facts(
    db: AsyncSession = Depends(deps.get_db),
    med_rep_id: int = None,
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    return await crud_sales.get_facts(db, med_rep_id=med_rep_id)

@router.get("/doctor-facts/", response_model=List[DoctorFactAssignmentSchema])
async def read_doctor_facts(
    db: AsyncSession = Depends(deps.get_db),
    skip: int = 0,
    limit: int = 100,
    med_rep_id: int = None,
    doctor_id: int = None,
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    return await crud_sales.get_doctor_fact_assignments(
        db, skip=skip, limit=limit, med_rep_id=med_rep_id, doctor_id=doctor_id
    )

@router.post("/doctor-facts/", response_model=DoctorFactAssignmentSchema)
async def create_doctor_fact(
    *,
    db: AsyncSession = Depends(deps.get_db),
    fact_in: DoctorFactAssignmentCreate,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    fact = await crud_sales.create_doctor_fact_assignment(db, obj_in=fact_in)
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "CREATE", "DoctorFact", fact.id,
        f"Факт врача прикреплен: {fact.quantity} шт.",
        request
    )
    return fact

@router.delete("/doctor-facts/{id}")
async def delete_doctor_fact(
    id: int,
    id_type: str = "fact",
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    """
    Удаление факта и отмена бонуса. id_type may be 'fact' or 'ledger'.
    """
    from app.services.finance_service import FinancialService
    return await FinancialService.delete_doctor_fact_assignment(db, id, id_type)

# Bonus Payments
@router.get("/bonus-payments/", response_model=List[BonusPaymentSchema])
async def read_bonus_payments(
    db: AsyncSession = Depends(deps.get_db),
    skip: int = 0,
    limit: int = 100,
    med_rep_id: int = None,
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    return await crud_sales.get_bonus_payments(
        db, skip=skip, limit=limit, med_rep_id=med_rep_id
    )

@router.post("/bonus-payments/", response_model=BonusPaymentSchema)
async def create_bonus_payment(
    *,
    db: AsyncSession = Depends(deps.get_db),
    payment_in: BonusPaymentCreate,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    allowed_roles = {UserRole.INVESTOR, UserRole.DEPUTY_DIRECTOR, UserRole.DIRECTOR, UserRole.ADMIN}
    if current_user.role not in allowed_roles:
        raise HTTPException(
            status_code=403,
            detail="Only Deputy Director or higher can record bonus payments"
        )
    payment = await crud_sales.create_bonus_payment(db, obj_in=payment_in)
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "CREATE", "BonusPayment", payment.id,
        f"Бонус выплачен: {payment.amount:,.0f} UZS",
        request
    )
    return payment

@router.patch("/bonus-payments/{payment_id}/", response_model=BonusPaymentSchema)
async def update_bonus_payment(
    *,
    db: AsyncSession = Depends(deps.get_db),
    payment_id: int,
    payment_in: BonusPaymentUpdate,
    current_user: User = Depends(deps.get_current_user),
    request: Request,
) -> Any:
    allowed_roles = {UserRole.INVESTOR, UserRole.DEPUTY_DIRECTOR, UserRole.DIRECTOR, UserRole.ADMIN}
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    result = await crud_sales.update_bonus_payment(db, payment_id=payment_id, obj_in=payment_in)
    if not result:
        raise HTTPException(status_code=404, detail="Bonus payment not found")
    
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "UPDATE", "BonusPayment", result.id,
        f"Выплата бонуса изменена: ID {payment_id}",
        request
    )
    return result

@router.get("/reservations/{id}/export")
async def export_reservation_excel(
    *,
    db: AsyncSession = Depends(deps.get_db),
    id: int,
    current_user: User = Depends(deps.get_current_user),
):
    try:
        reservation = await crud_sales.get_reservation(db, id=id)
        if not reservation:
            raise HTTPException(status_code=404, detail="Reservation not found")
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Фактура"
        
        # Yellow fill
        fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_blue = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        fill_pink = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
        
        font_bold = Font(bold=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 1. Харажатлардан кейин Фойда/Зарар (Shifting to column 4 - Column D)
        ws.cell(row=1, column=4, value="Харажатлардан кейин Фойда/Зарар").font = font_bold
        ws.cell(row=2, column=4, value="Клинт бонуси")
        ws.cell(row=2, column=5, value=0).fill = fill_yellow
        ws.cell(row=3, column=4, value="Менеджерлар бонуси")
        ws.cell(row=4, column=4, value="Прямой").font = font_bold
        ws.cell(row=4, column=5).fill = fill_yellow
        ws.cell(row=5, column=4, value="Доставка")
        ws.cell(row=5, column=5).fill = fill_yellow
        
        org_name = reservation.med_org.name if reservation.med_org else (reservation.customer_name or "N/A")
        org_inn = reservation.med_org.inn if reservation.med_org and reservation.med_org.inn else ""
        
        ws.cell(row=6, column=4, value="КОРХОНА НОМИ").fill = fill_blue
        ws.cell(row=6, column=5, value=org_name).fill = fill_blue
        ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=8)
        
        ws.cell(row=7, column=4, value="ИНН").fill = fill_blue
        ws.cell(row=7, column=5, value=org_inn).fill = fill_blue
        ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=8)
        
        headers = ["№", "Наименование", "Кол-во", "Цена", "Завода келишилган", "Сумма"]
        for i, header in enumerate(headers):
            col_idx = i + 3 # Start from C
            cell = ws.cell(row=8, column=col_idx, value=header)
            cell.border = border_thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if col_idx == 7: # "Завода келишилган"
                cell.fill = fill_pink
            else:
                cell.fill = fill_blue
            
        # Column widths (Shifting A->C, B->D, etc.)
        ws.column_dimensions['C'].width = 5
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 15
        
        subtotal_plain = 0.0
        row_idx = 9
        for idx, item in enumerate(reservation.items, 1):
            actual_qty = (item.quantity or 0) - (item.returned_quantity or 0)
            if actual_qty <= 0:
                continue
                
            item_price = item.price or 0.0
            # Plain amount for the row (qty * price)
            plain_amount = actual_qty * item_price
            subtotal_plain += plain_amount
            
            ws.cell(row=row_idx, column=3, value=idx).border = border_thin
            
            cell_name = ws.cell(row=row_idx, column=4, value=item.product.name if item.product else "")
            cell_name.border = border_thin
            cell_name.fill = fill_yellow
            
            cell_qty = ws.cell(row=row_idx, column=5, value=actual_qty)
            cell_qty.border = border_thin
            cell_qty.fill = fill_yellow
            cell_qty.alignment = Alignment(horizontal='center')
            
            cell_price = ws.cell(row=row_idx, column=6, value=item_price)
            cell_price.border = border_thin
            cell_price.fill = fill_yellow
            cell_price.alignment = Alignment(horizontal='center')
            
            # Завода келишилган (production price)
            prod_price = item.product.production_price if item.product else 0
            cell_prod = ws.cell(row=row_idx, column=7, value=prod_price)
            cell_prod.border = border_thin
            cell_prod.fill = fill_pink
            cell_prod.alignment = Alignment(horizontal='center')
            
            cell_sum = ws.cell(row=row_idx, column=8, value=plain_amount)
            cell_sum.border = border_thin
            cell_sum.fill = fill_yellow
            cell_sum.alignment = Alignment(horizontal='right')
            row_idx += 1
            
        # Add 4 extra empty rows (as requested)
        for _ in range(4):
            for col_idx in range(3, 9): # Columns C to H
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border_thin
            row_idx += 1
            
        # Totals
        discount_val = 0.0
        # If any item has a discount, we use it (or reservation-level if we had one)
        if reservation.items and len(reservation.items) > 0:
            discount_val = next((it.discount_percent for it in reservation.items if it.discount_percent), 0.0)
            
        discounted_total = subtotal_plain * (1 - discount_val / 100.0)
        nds_percent = reservation.nds_percent if reservation.nds_percent is not None else 12.0
        nds_total = discounted_total * (1 + nds_percent / 100.0)
        
        cell_sub = ws.cell(row=row_idx, column=8, value=subtotal_plain)
        cell_sub.font = font_bold
        cell_sub.alignment = Alignment(horizontal='right')
        
        label_disc = ws.cell(row=row_idx + 1, column=7, value=f"{discount_val}% скидка билан")
        label_disc.alignment = Alignment(horizontal='right')
        cell_disc = ws.cell(row=row_idx + 1, column=8, value=discounted_total)
        cell_disc.font = font_bold
        cell_disc.alignment = Alignment(horizontal='right')
        
        label_nds = ws.cell(row=row_idx + 2, column=7, value=f"{nds_percent}% ндс билан")
        label_nds.alignment = Alignment(horizontal='right')
        cell_nds = ws.cell(row=row_idx + 2, column=8, value=nds_total)
        cell_nds.font = font_bold
        cell_nds.alignment = Alignment(horizontal='right')
        
        org_name = (reservation.med_org.name if reservation.med_org else reservation.customer_name) or "N/A"
        org_inn = reservation.med_org.inn if reservation.med_org and reservation.med_org.inn else "no_inn"
        realization_date = reservation.invoice.realization_date if reservation.invoice and reservation.invoice.realization_date else reservation.date
        date_str = realization_date.strftime("%d.%m.%Y") if realization_date else "no_date"
        
        # Sanitize for filename: remove truly illegal chars like / \ : * ? " < > |
        # We preserve spaces, dots, and Unicode (Cyrillic) characters
        illegal_chars = '/\\:*?"<>|'
        safe_org_name = "".join([c for c in org_name if c not in illegal_chars]).strip()
        filename = f"{safe_org_name}_{org_inn}_{date_str}.xlsx"
        encoded_filename = urllib.parse.quote(filename)
        
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        return StreamingResponse(
            stream, 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Access-Control-Expose-Headers": "Content-Disposition"
            }
        )
    except Exception as e:
        logger.error(f"Error exporting reservation {id}: {traceback.format_exc()}")
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

# MedRep Bonus Balance System

@router.get("/bonuses/history/{med_rep_id}")
async def get_bonus_history(
    med_rep_id: int,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    """
    Alias for get_medrep_bonus_balance to match frontend expectations.
    """
    return await get_medrep_bonus_balance(med_rep_id=med_rep_id, db=db, current_user=current_user)

@router.get("/bonus-balance/")
async def get_medrep_bonus_balance(
    med_rep_id: Optional[int] = None,
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    """
    Get the current bonus balance and transaction history for a MedRep.
    If med_rep_id is not provided, uses current_user.id.
    Managers and Directors can view any MedRep's balance.
    """
    try:
        target_id = med_rep_id if med_rep_id is not None else current_user.id
        
        # Permission check
        # MedReps can only see their own balance
        if current_user.role == UserRole.MED_REP and target_id != current_user.id:
            raise HTTPException(status_code=403, detail="You can only view your own balance")
        
        # Non-MedReps must be managers/directors to view balances
        allowed_roles = [UserRole.DIRECTOR, UserRole.DEPUTY_DIRECTOR, UserRole.PRODUCT_MANAGER, UserRole.REGIONAL_MANAGER, UserRole.ADMIN]
        if current_user.role not in allowed_roles and current_user.role != UserRole.MED_REP:
            raise HTTPException(status_code=403, detail="Access denied")
        
        from app.services.finance_service import FinancialService
        from app.models.ledger import BonusLedger, LedgerType
        from app.models.sales import Reservation, ReservationItem, Invoice
        # Calculate usable balance (Paid accruals - offsets)
        balance = await FinancialService.get_medrep_bonus_balance(db, target_id)
        

        
        all_entries_res = await db.execute(select(BonusLedger).where(BonusLedger.user_id == target_id))
        all_entries = all_entries_res.scalars().all()
        
        total_accrued = 0.0
        total_paid = 0.0
        total_allocated = 0.0
        
        for e in all_entries:
            if e.ledger_type == LedgerType.ACCRUAL:
                total_accrued += e.amount
                if e.is_paid:
                    total_paid += e.amount
            elif e.ledger_type == LedgerType.OFFSET:
                total_allocated += abs(e.amount)
        
        # Get history (all transactions)
        query = select(BonusLedger).options(
            selectinload(BonusLedger.doctor),
            selectinload(BonusLedger.product),
            selectinload(BonusLedger.payment),
            selectinload(BonusLedger.invoice_item).selectinload(ReservationItem.reservation).selectinload(Reservation.invoice)
        ).where(BonusLedger.user_id == target_id)
        
        # Filter by month/year if provided
        if month and year:
            query = query.where(BonusLedger.target_month == month, BonusLedger.target_year == year)
        else:
            # Default: last 30 days
            from datetime import datetime, timedelta
            thirty_days_ago = datetime.utcnow() - timedelta(days=30)
            query = query.where(BonusLedger.created_at >= thirty_days_ago)
            
        query = query.order_by(BonusLedger.created_at.desc())
        
        result = await db.execute(query)
        history = result.scalars().all()
        
        import re
        from app.models.sales import Invoice
        invoice_ids = set()
        for h in history:
            if getattr(h, 'payment', None) and getattr(h.payment, 'invoice_id', None):
                invoice_ids.add(h.payment.invoice_id)
            elif h.notes:
                match = re.search(r'#(?:СФ-)?(\d+)', h.notes)
                if match:
                    invoice_ids.add(int(match.group(1)))
                    
        invoice_to_reservation = {}
        if invoice_ids:
            inv_query = select(Invoice.id, Invoice.reservation_id).where(Invoice.id.in_(invoice_ids))
            inv_result = await db.execute(inv_query)
            for inv_id, res_id in inv_result.all():
                invoice_to_reservation[inv_id] = res_id
        
        # Map to dictionaries to avoid JSON serialization errors
        history_data = []
        for h in history:
            inv_id = None
            res_id = None
            
            if getattr(h, 'invoice_item', None) and getattr(h.invoice_item, 'reservation', None):
                res_id = h.invoice_item.reservation.id
                if getattr(h.invoice_item.reservation, 'invoice', None):
                    inv_id = h.invoice_item.reservation.invoice.id
                    
            if not res_id:
                if getattr(h, 'payment', None) and getattr(h.payment, 'invoice_id', None):
                    inv_id = h.payment.invoice_id
                elif h.notes:
                    match = re.search(r'#(?:СФ-)?(\d+)', h.notes)
                    if match:
                        inv_id = int(match.group(1))
                
                if inv_id and inv_id in invoice_to_reservation:
                    res_id = invoice_to_reservation[inv_id]

            history_data.append({
                "id": h.id,
                "amount": h.amount,
                "ledger_type": h.ledger_type,
                "created_at": h.created_at.isoformat(),
                "notes": h.notes,
                "invoice_id": inv_id,
                "reservation_id": res_id,
                "target_month": h.target_month,
                "target_year": h.target_year,
                "is_paid": h.is_paid,
                "doctor": {"id": h.doctor.id, "full_name": h.doctor.full_name} if h.doctor else None,
                "product": {"id": h.product.id, "name": h.product.name} if h.product else None,
                "payment_amount": h.payment.amount if h.payment else None,
                "payment_type": h.payment.payment_type if h.payment else None,
            })
        
        return {
            "balance": balance,
            "total_accrued": total_accrued,
            "total_paid": total_paid,
            "total_allocated": total_allocated,
            "history": history_data
        }
    except Exception as e:
        import traceback
        logger.error(f"Error in get_medrep_bonus_balance: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
@router.post("/allocate-bonus/")
async def allocate_bonus(
    *,
    db: AsyncSession = Depends(deps.get_db),
    alloc_in: BonusAllocationCreate,
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    """
    Любой авторизованный пользователь может распределять бонусы врачам.
    """
    try:
        from app.services.finance_service import FinancialService
        
        # Если указан med_rep_id — дебетуем баланс указанного медпреда (для admin/director), иначе текущего пользователя
        effective_med_rep_id = alloc_in.med_rep_id if alloc_in.med_rep_id else current_user.id
        
        result = await FinancialService.allocate_bonus(
            db=db,
            med_rep_id=effective_med_rep_id,
            doctor_id=alloc_in.doctor_id,
            product_id=alloc_in.product_id,
            quantity=alloc_in.quantity,
            amount_per_unit=alloc_in.amount_per_unit,
            target_month=alloc_in.target_month,
            target_year=alloc_in.target_year,
            notes=alloc_in.notes
        )
        
        return result
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        raise HTTPException(status_code=500, detail=traceback.format_exc())

# ==========================================
# Admin Bonus Approvals (Director/Admin)
# ==========================================

from pydantic import BaseModel

class BonusSummary(BaseModel):
    med_rep_id: int
    med_rep_name: str
    accrued: float # Начислено всего (Факт)
    paid: float    # Выплачено директором
    remainder: float # Остаток к выплате
    allocated: float # Распределено врачам
    predinvest: float # Аванс (Предынвест)
    realization: float = 0.0
    postupleniya: float = 0.0
    debitorka: float = 0.0
    has_overdue_bonus: bool = False

class GlobalStats(BaseModel):
    realization: float = 0.0
    postupleniya: float = 0.0
    debitorka: float = 0.0

class AdminBonusSummaryResponse(BaseModel):
    summaries: List[BonusSummary]
    global_stats: GlobalStats

class BonusPayRequest(BaseModel):
    med_rep_id: int
    amount_to_pay: float # How much of the remainder to pay now

@router.get("/admin/bonuses/summary", response_model=AdminBonusSummaryResponse)
async def get_admin_bonus_summary(
    month: int = None,
    year: int = None,
    product_id: int = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user),
) -> Any:
    """
    Returns a summary of bonuses for all MedReps.
    Only accessible by Director, Deputy Director, Admin.
    """
    allowed_roles = {UserRole.INVESTOR, UserRole.DEPUTY_DIRECTOR, UserRole.DIRECTOR, UserRole.ADMIN}
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Insufficient permissions")

    from app.models.user import User
    from app.models.ledger import BonusLedger, LedgerType
    from app.models.sales import Invoice, Payment, Reservation, ReservationItem, InvoiceStatus
    from datetime import datetime, timedelta
    from sqlalchemy import func, and_, or_
    
    # Get all medreps
    medreps_result = await db.execute(select(User).where(User.role == UserRole.MED_REP, User.is_active == True))
    medreps = medreps_result.scalars().all()
    rep_ids = [r.id for r in medreps]
    
    start_date = None
    end_date = None
    if month and year:
        start_date = datetime(year, month, 1)
        end_date = (datetime(year, month + 1, 1) if month < 12 else datetime(year + 1, 1, 1))
    elif year:
        start_date = datetime(year, 1, 1)
        end_date = datetime(year + 1, 1, 1)
        
    realization_map = {}
    postupleniya_map = {}
    debitorka_map = {}
    
    if rep_ids:
        for rep in medreps:
            # REALIZATION
            real_q = select(
                func.coalesce(func.sum(ReservationItem.quantity * ReservationItem.price), 0.0)
            ).select_from(Invoice)\
             .join(Reservation, Invoice.reservation_id == Reservation.id)\
             .join(ReservationItem, Reservation.id == ReservationItem.reservation_id)\
             .outerjoin(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id)\
             .where(Invoice.status != InvoiceStatus.CANCELLED)\
             .where(or_(
                 Reservation.created_by_id == rep.id, 
                 MedicalOrganization.assigned_reps.any(User.id == rep.id)
             ))
             
            if start_date and end_date: real_q = real_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
            if product_id: real_q = real_q.where(ReservationItem.product_id == product_id)
            realization_map[rep.id] = float((await db.execute(real_q)).scalar() or 0.0)
                
            # POSTUPLENIYA
            pay_q = select(
                func.coalesce(func.sum(Payment.amount * (ReservationItem.quantity * ReservationItem.price / Invoice.total_amount)), 0.0)
            ).select_from(Payment)\
             .join(Invoice, Payment.invoice_id == Invoice.id)\
             .join(Reservation, Invoice.reservation_id == Reservation.id)\
             .join(ReservationItem, Reservation.id == ReservationItem.reservation_id)\
             .outerjoin(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id)\
             .where(Invoice.total_amount > 0)\
             .where(or_(
                 Reservation.created_by_id == rep.id, 
                 MedicalOrganization.assigned_reps.any(User.id == rep.id)
             ))
             
            if start_date and end_date: pay_q = pay_q.where(and_(Payment.date >= start_date, Payment.date < end_date))
            if product_id: pay_q = pay_q.where(ReservationItem.product_id == product_id)
            postupleniya_map[rep.id] = float((await db.execute(pay_q)).scalar() or 0.0)
                
            # DEBITORKA
            debt_q = select(
                func.coalesce(func.sum(
                    (ReservationItem.quantity * ReservationItem.price) -
                    (ReservationItem.quantity * ReservationItem.price / Invoice.total_amount * Invoice.paid_amount)
                ), 0.0)
            ).select_from(Invoice)\
             .join(Reservation, Invoice.reservation_id == Reservation.id)\
             .join(ReservationItem, Reservation.id == ReservationItem.reservation_id)\
             .outerjoin(MedicalOrganization, Reservation.med_org_id == MedicalOrganization.id)\
             .where(Invoice.status.in_([InvoiceStatus.UNPAID, InvoiceStatus.PARTIAL, InvoiceStatus.APPROVED]), Invoice.total_amount > 0)\
             .where(or_(
                 Reservation.created_by_id == rep.id, 
                 MedicalOrganization.assigned_reps.any(User.id == rep.id)
             ))
             
            if start_date and end_date: debt_q = debt_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
            if product_id: debt_q = debt_q.where(ReservationItem.product_id == product_id)
            debitorka_map[rep.id] = float((await db.execute(debt_q)).scalar() or 0.0)

    # GLOBAL AGGREGATES (Independent of MedReps)
    # Realization
    g_real_q = select(func.coalesce(func.sum(ReservationItem.quantity * ReservationItem.price), 0.0)).select_from(Invoice)\
        .join(Reservation, Invoice.reservation_id == Reservation.id)\
        .join(ReservationItem, Reservation.id == ReservationItem.reservation_id)\
        .where(Invoice.status != InvoiceStatus.CANCELLED)
    if start_date and end_date: g_real_q = g_real_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if product_id: g_real_q = g_real_q.where(ReservationItem.product_id == product_id)
    global_realization = float((await db.execute(g_real_q)).scalar() or 0.0)

    # Postupleniya
    g_pay_q = select(func.coalesce(func.sum(Payment.amount * (ReservationItem.quantity * ReservationItem.price / Invoice.total_amount)), 0.0)).select_from(Payment)\
        .join(Invoice, Payment.invoice_id == Invoice.id)\
        .join(Reservation, Invoice.reservation_id == Reservation.id)\
        .join(ReservationItem, Reservation.id == ReservationItem.reservation_id)\
        .where(Invoice.total_amount > 0)
    if start_date and end_date: g_pay_q = g_pay_q.where(and_(Payment.date >= start_date, Payment.date < end_date))
    if product_id: g_pay_q = g_pay_q.where(ReservationItem.product_id == product_id)
    global_postupleniya = float((await db.execute(g_pay_q)).scalar() or 0.0)

    # Debitorka
    g_debt_q = select(func.coalesce(func.sum(
        (ReservationItem.quantity * ReservationItem.price) -
        (ReservationItem.quantity * ReservationItem.price / Invoice.total_amount * Invoice.paid_amount)
    ), 0.0)).select_from(Invoice)\
        .join(Reservation, Invoice.reservation_id == Reservation.id)\
        .join(ReservationItem, Reservation.id == ReservationItem.reservation_id)\
        .where(Invoice.status.in_([InvoiceStatus.UNPAID, InvoiceStatus.PARTIAL, InvoiceStatus.APPROVED]), Invoice.total_amount > 0)
    if start_date and end_date: g_debt_q = g_debt_q.where(and_(Invoice.date >= start_date, Invoice.date < end_date))
    if product_id: g_debt_q = g_debt_q.where(ReservationItem.product_id == product_id)
    global_debitorka = float((await db.execute(g_debt_q)).scalar() or 0.0)

    summaries = []
    for rep in medreps:
        q = select(BonusLedger).where(BonusLedger.user_id == rep.id)
        if start_date and end_date: q = q.where(and_(BonusLedger.created_at >= start_date, BonusLedger.created_at < end_date))
        if product_id: q = q.where(BonusLedger.product_id == product_id)
        
        ledger_res = await db.execute(q)
        entries = ledger_res.scalars().all()
        
        accrued = 0.0
        paid = 0.0
        allocated = 0.0
        predinvest = 0.0
        
        for e in entries:
            if e.ledger_type == LedgerType.ACCRUAL:
                if e.notes == "Аванс (Предынвест)":
                    predinvest += e.amount
                    paid += e.amount
                else:
                    accrued += e.amount
                    if e.is_paid:
                        paid += e.amount
            elif e.ledger_type == LedgerType.OFFSET:
                allocated += abs(e.amount)
                
        remainder = max(0.0, accrued - paid)
        
        overdue_q = select(BonusLedger.id).where(
            and_(
                BonusLedger.user_id == rep.id,
                BonusLedger.ledger_type == LedgerType.ACCRUAL,
                BonusLedger.is_paid == False,
                or_(BonusLedger.notes != "Аванс (Предынвест)", BonusLedger.notes.is_(None)),
                BonusLedger.created_at < datetime.utcnow() - timedelta(days=15)
            )
        ).limit(1)
        overdue_res = await db.execute(overdue_q)
        has_overdue = overdue_res.scalar_one_or_none() is not None
        
        summaries.append(BonusSummary(
            med_rep_id=rep.id,
            med_rep_name=rep.full_name,
            accrued=accrued,
            paid=paid,
            remainder=remainder,
            allocated=allocated,
            predinvest=predinvest,
            realization=realization_map.get(rep.id, 0.0),
            postupleniya=postupleniya_map.get(rep.id, 0.0),
            debitorka=debitorka_map.get(rep.id, 0.0),
            has_overdue_bonus=has_overdue
        ))
        
    return AdminBonusSummaryResponse(
        summaries=summaries,
        global_stats=GlobalStats(
            realization=global_realization,
            postupleniya=global_postupleniya,
            debitorka=global_debitorka
        )
    )

@router.post("/admin/bonuses/pay")
async def pay_medrep_bonus(
    request_data: BonusPayRequest,
    request: Request,
    db: AsyncSession = Depends(deps.get_db),
    current_user: User = Depends(deps.get_current_user)
) -> Any:
    """
    Marks unpaid ACCRUAL records as paid up to the requested amount.
    """
    allowed_roles = {UserRole.INVESTOR, UserRole.DEPUTY_DIRECTOR, UserRole.DIRECTOR, UserRole.ADMIN}
    if current_user.role not in allowed_roles:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
        
    from app.models.ledger import BonusLedger, LedgerType
    
    # Get unpaid accruals for this medrep, sorted by oldest first
    query = select(BonusLedger).where(
        BonusLedger.user_id == request_data.med_rep_id,
        BonusLedger.ledger_type == LedgerType.ACCRUAL,
        BonusLedger.is_paid == False
    ).order_by(BonusLedger.id.asc())
    
    result = await db.execute(query)
    unpaid_entries = result.scalars().all()
    
    amount_remaining_to_pay = request_data.amount_to_pay
    actual_paid = 0.0
    
    for entry in unpaid_entries:
        if amount_remaining_to_pay <= 0:
            break
            
        # We pay the entry as long as it's <= our remaining amount to pay
        if entry.amount <= amount_remaining_to_pay:
            entry.is_paid = True
            amount_remaining_to_pay -= entry.amount
            actual_paid += entry.amount
        else:
            # We pay a portion of it by splitting it? 
            # Actually, standard behavior here is to just mark it as paid if amount covers it.
            # If the admin wants to pay partial, usually we don't handle partial invoice payments this way.
            # But the original code just skipped partials. We'll leave it as is to avoid breaking existing logic.
            pass
            
    # If there is STILL remaining money to pay, it's an advance payment (Predinvest). 
    # Create a new BonusLedger entry for this excess.
    if amount_remaining_to_pay > 0:
        from datetime import datetime
        now = datetime.utcnow()
        predinvest_entry = BonusLedger(
            user_id=request_data.med_rep_id,
            amount=amount_remaining_to_pay,
            ledger_type=LedgerType.ACCRUAL,
            is_paid=True, # It is immediately paid out
            target_month=now.month,
            target_year=now.year,
            notes=f"Аванс (Предынвест)"
        )
        db.add(predinvest_entry)
        actual_paid += amount_remaining_to_pay

    await db.commit()
    
    from app.services.audit_service import log_action
    await log_action(
        db, current_user, "UPDATE", "BonusLedger", request_data.med_rep_id,
        f"Выплачен бонус МП: {actual_paid:,.0f} UZS",
        request
    )
        
    return {"message": "Успешно выплачено", "paid_amount": actual_paid}
